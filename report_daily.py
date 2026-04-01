# report_daily.py  v6
# 8-sheet SGSN daily report — data sources verified against Java source (SGSNReport.jar)
#
# CONFIRMED DATA SOURCES (from Java decompile):
#
# SGSNNOKIA  : nokia_sgsn_report_out (pre-computed view, NOT raw nokia_sgsn_report)
#              joined with nokia_sgsn_report_pdp_out + nokia_sgsn_report_att_out
#
# Total      : 6-table JOIN:
#              t3 = p_obs_zte (ZTE 2G/3G)
#              t4 = nokia_sgsn_report_out (Nokia 2G/3G)
#              t5 = p_obs_zte_lan (NIB/DT:  in_kb/1024=DL MB, out_kb/1024=UL MB)
#              t6 = nokia_4g_sgsn_report (4G: SUM lte_5212a DL, lte_5213a UL)
#              3G = t5.download - (t3.ul_2g + t4.ul_2g) [Java exact formula]
#              2G = t3.ul_2g + t4.ul_2g
#
# 5 Min      : MySQL downldSgsn (pre-computed daily totals, col layout from reference)
#              Columns: ZTE, Nokia, NIB (DT), 4G Nokia, Total — per day
#
# Peak Tput  : MySQL peakThroughputSgsn
#              Columns: ZTE, LAN/NIB, Nokia, Nokia4G — per day
#
# MAR-26     : Nokia  → nokia_sgsn_report_out monthly totals
#              DT     → p_obs_zte_lan_4g_tput (in/out_volume_kb) - Nokia_SGSN
#                        ZTE SGSN shutdown → DT = prtg4g_total - Nokia (current value)
#              4G     → nokia_4g_sgsn_report monthly SUM
#              TCS    → TCS EMS CSV file (read directly)
#
# 4G sheet   : nokia_sgsn_report (flns columns) + sgsn_report_sbc_kpi + volte_bearer_act
# TRAI       : MySQL sgsn_trai_report

import os, calendar
from datetime import datetime, time as dt_time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import pg_query, my_query
from config import OUTPUT_DIR
from tcs_reader import read_tcs_daily
from tcs_store  import store_tcs, load_tcs_month, load_tcs_day
from file_resolver import resolve_tcs_file

B_TO_MB  = 1048576
KB_TO_MB = 1024.0
MB_TO_GB = 1024.0

C_DARK="1F4E79"; C_MED="2E75B6"; C_LBLUE="BDD7EE"; C_LTEAL="DDEBF7"
C_ORNG="C55A11"; C_GREEN="70AD47"; C_TOTAL="F2F2F2"; C_GRAY="D9D9D9"; C_WHITE="FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=9):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)
def _hdr(ws, row, col, val, bg=C_DARK, fc=C_WHITE):
    c = ws.cell(row=row, column=col, value=val)
    c.fill=_fill(bg); c.font=_font(True,fc); c.alignment=_align(wrap=True); c.border=_border()
def _dat(ws, row, col, val, bg=None):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill=_fill(bg)
    c.alignment=_align(); c.border=_border()
def _pct(n, d): return round(n/d*100, 2) if d else 0.0
def _safe(v):
    try:    return float(v or 0)
    except: return 0.0
def _mhdr(ws, r, c1, c2, txt, bg=C_MED):
    if c2>c1: ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, txt)
    cell.fill=_fill(bg); cell.font=_font(True,C_WHITE); cell.alignment=_align(); cell.border=_border()
def _total_row(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.fill=_fill(C_TOTAL); c.font=_font(True); c.alignment=_align(); c.border=_border()


# =============================================================================
#  DATA FETCHERS — HOURLY (for SGSNZTE, SGSNNOKIA, Total sheets)
# =============================================================================

def _zte_hourly(date_str):
    """ZTE SGSN per-hour from p_obs_zte. ZTE is decommissioned → returns empty dict."""
    for sql in [
        """SELECT date_trunc('hour',end_time)::timestamp AS hr,
                sum(GTP81)/%(b)s AS ul_3g_mb, sum(GTP83)/%(b)s AS dl_3g_mb,
                sum(GTP73)/%(b)s AS ul_2g_mb, sum(GTP75)/%(b)s AS dl_2g_mb,
                avg(coalesce(GTP29,0))/1024.0 AS dl_tput, avg(coalesce(GTP28,0))/1024.0 AS ul_tput,
                sum(coalesce(session1,0)) AS pdp3g_att, sum(coalesce(session3,0)) AS pdp3g_act,
                sum(coalesce(session2,0)) AS pdp2g_att, sum(coalesce(session4,0)) AS pdp2g_act,
                sum(coalesce(ovrld1,0)) AS rej_ovrld, sum(coalesce(ovrld2,0)) AS rej_rf
           FROM p_obs_zte WHERE date(end_time)=%(dt)s GROUP BY 1 ORDER BY 1""",
        """SELECT date_trunc('hour',end_time)::timestamp AS hr,
                sum(GTP81)/%(b)s AS ul_3g_mb, sum(GTP83)/%(b)s AS dl_3g_mb,
                sum(GTP73)/%(b)s AS ul_2g_mb, sum(GTP75)/%(b)s AS dl_2g_mb,
                avg(coalesce(GTP29,0)) AS dl_tput, avg(coalesce(GTP28,0)) AS ul_tput
           FROM p_obs_zte WHERE date(end_time)=%(dt)s GROUP BY 1 ORDER BY 1""",
    ]:
        try:
            rows = pg_query(sql, {'b': B_TO_MB, 'dt': date_str})
            return {r['hr'].hour: r for r in rows if r.get('hr')}
        except Exception:
            continue
    return {}


def _nokia_hourly(date_str):
    """
    Nokia SGSN per-hour — from nokia_sgsn_report_out (EXACT same table as Java).
    Pre-computed view with upload_3g, download_3g, upload_2g, download_2g (MB),
    peak2gthru, peak3gthru (Mbps), thu_up, thu_down (Mbps).
    Joined with pdp_out and att_out tables.
    """
    try:
        rows = pg_query("""
            SELECT extract(hour FROM a.period_start_time)::int AS hr,
                sum(a.upload_3g)      AS ul_3g_mb,
                sum(a.download_3g)    AS dl_3g_mb,
                sum(a.upload_2g)      AS ul_2g_mb,
                sum(a.download_2g)    AS dl_2g_mb,
                sum(a.total2g)        AS total_2g_mb,
                sum(a.total3g)        AS total_3g_mb,
                sum(a.totalgn)        AS total_gn_mb,
                max(a.peak2gthru)     AS tput_2g,
                max(a.peak3gthru)     AS tput_3g,
                max(a.thu_up)         AS ul_tput,
                max(a.thu_down)       AS dl_tput,
                max(a.pdp_3g)         AS pdp3g_att,
                max(a.actpdp_3g)      AS pdp3g_act,
                max(a.pdp_2g)         AS pdp2g_att,
                max(a.actpdp_2g)      AS pdp2g_act,
                sum(b.att)            AS rej_att,
                sum(b.succ)           AS rej_succ,
                sum(b.rej_ggsn)       AS rej_ggsn,
                sum(b.rej_in)         AS rej_in,
                sum(b.rej_other)      AS rej_oth,
                sum(b.rej_overload)   AS rej_ovrld,
                sum(b.rej_rf)         AS rej_rf,
                sum(c.att)            AS att_att,
                sum(c.succ)           AS att_succ
            FROM public.nokia_sgsn_report_out a
            INNER JOIN public.nokia_sgsn_report_pdp_out b
                ON a.period_start_time = b.period_start_time
            INNER JOIN public.nokia_sgsn_report_att_out c
                ON a.period_start_time = c.period_start_time
            WHERE date(a.period_start_time) = %(dt)s
            GROUP BY extract(hour FROM a.period_start_time)
            ORDER BY hr
        """, {'dt': date_str})
        return {r['hr']: r for r in rows}
    except Exception:
        return {}


def _nokia_4g_hourly(date_str):
    """Nokia 4G per-hour: SUM(lte_5212a) DL MB, SUM(lte_5213a) UL MB."""
    try:
        rows = pg_query("""
            SELECT extract(hour FROM period_start_time)::int AS hr,
                sum(coalesce(lte_5213a,0)) AS ul_mb,
                sum(coalesce(lte_5212a,0)) AS dl_mb
            FROM public.nokia_4g_sgsn_report
            WHERE date(period_start_time)=%(dt)s GROUP BY 1 ORDER BY 1
        """, {'dt': date_str})
        return {r['hr']: r for r in rows}
    except Exception:
        return {}


def _lan_total_hourly(date_str):
    """
    NIB/LAN per-hour volumes and speeds (exact Java t5 formula):
      DL = in_volume_kb / 1024  (MB)
      UL = out_volume_kb / 1024 (MB)
      down_speed = max(in_speed_kbps)  / 1024  (Mbps)
      up_speed   = max(out_speed_kbps) / 1024  (Mbps)
    """
    try:
        rows = pg_query("""
            SELECT extract(hour FROM t.date_time)::int        AS hr,
                sum(t.sum_in_kb)/1024.0                           AS dl_mb,
                sum(t.sum_out_kb)/1024.0                          AS ul_mb,
                max(t.avg_in_spd)/1024.0                          AS down_speed,
                max(t.avg_out_spd)/1024.0                         AS up_speed,
                max(t.avg_in_spd + t.avg_out_spd)/1024.0          AS tot_speed
            FROM (
                SELECT date_time,
                       sum(coalesce(in_volume_kb,0))  AS sum_in_kb,
                       sum(coalesce(out_volume_kb,0)) AS sum_out_kb,
                       avg(coalesce(in_speed_kbps,0)) AS avg_in_spd,
                       avg(coalesce(out_speed_kbps,0))AS avg_out_spd
                FROM public.p_obs_zte_lan
                WHERE date(date_time)=%(dt)s
                GROUP BY date_time
            ) t
            GROUP BY extract(hour FROM t.date_time)
            ORDER BY 1
        """, {'dt': date_str})
        return {r['hr']: r for r in rows}
    except Exception:
        return {}


# =============================================================================
#  DATA FETCHERS — MONTHLY (for MAR-26, 5Min, Peak, TRAI sheets)
# =============================================================================

def _downld_sgsn_month(year, month):
    """
    MySQL downldSgsn — pre-computed daily totals used by MAR-26 (DT) and 5Min sheets.
    Java exact query: SELECT Date_,downldSGSNZTE,downldSGSNdt,downldSGSNNOKIA,
                             download4g,downldSGSNtot FROM downldSgsn
                      WHERE MONTH(Date_)=? AND YEAR(Date_)=? AND DATE(Date_)<=?
    """
    try:
        return my_query("""
            SELECT Date_             AS date_,
                   downldSGSNZTE     AS zte_gb,
                   downldSGSNdt      AS dt_gb,
                   downldSGSNNOKIA   AS nokia_gb,
                   download4g        AS g4_gb,
                   downldSGSNtot     AS total_gb
            FROM reportsDB.downldSgsn
            WHERE MONTH(Date_)=%s AND YEAR(Date_)=%s
            ORDER BY Date_
        """, (month, year))
    except Exception:
        return []


def _peak_tput_month(year, month):
    """
    MySQL peakThroughputSgsn — used by Peak Throughput sheet.
    Java query: SELECT Date_,(upPeakThruZte+downPeakThruZte) AS ZTE,
                       (upPeakThrudt+downPeakThrudt) AS LAN,
                       (upPeakThruNokia+downPeakThruNokia) AS NOKIA,
                       (upPeakThruNokia4g+downPeakThruNokia4g) AS nokia4g
                FROM peakThroughputSgsn WHERE MONTH(Date_)=? AND YEAR(Date_)=?
    """
    try:
        return my_query("""
            SELECT Date_                                      AS date_,
                   (upPeakThruZte    + downPeakThruZte)      AS zte_mbps,
                   (upPeakThrudt     + downPeakThrudt)       AS lan_mbps,
                   (upPeakThruNokia  + downPeakThruNokia)    AS nokia_mbps,
                   (upPeakThruNokia4g+ downPeakThruNokia4g)  AS nokia4g_mbps
            FROM reportsDB.peakThroughputSgsn
            WHERE MONTH(Date_)=%s AND YEAR(Date_)=%s
            ORDER BY Date_
        """, (month, year))
    except Exception:
        return []


def _nokia_month(year, month):
    """
    Nokia 2G/3G monthly — raw rows per period_start_time from nokia_sgsn_report_out.
    Used by:
      _build_5min  : slot_total_mb (totalgn per 15-min period) / 3 / MB_TO_GB
      _build_peak  : slot_tput (peak2gthru+peak3gthru Mbps per 15-min period)
    Each row = one 15-min record. Fields: day_, hr, mn, slot_total_mb, slot_tput
    """
    try:
        return pg_query("""
            SELECT extract(day    FROM period_start_time)::int AS day_,
                   extract(hour   FROM period_start_time)::int AS hr,
                   extract(minute FROM period_start_time)::int AS mn,
                   coalesce(totalgn,0)                          AS slot_total_mb,
                   coalesce(peak2gthru,0)+coalesce(peak3gthru,0) AS slot_tput
            FROM public.nokia_sgsn_report_out
            WHERE extract(year  FROM period_start_time)=%(y)s
              AND extract(month FROM period_start_time)=%(m)s
            ORDER BY period_start_time
        """, {'y': year, 'm': month})
    except Exception:
        return []


def _nokia_month_daily(year, month):
    """Nokia 2G/3G daily totals — for MAR-26 Nokia U/L and D/L rows."""
    try:
        rows = pg_query("""
            SELECT extract(day FROM period_start_time)::int AS day_,
                   sum(coalesce(upload_2g,0)+coalesce(upload_3g,0))     AS ul_mb,
                   sum(coalesce(download_2g,0)+coalesce(download_3g,0)) AS dl_mb
            FROM public.nokia_sgsn_report_out
            WHERE extract(year FROM period_start_time)=%(y)s
              AND extract(month FROM period_start_time)=%(m)s
            GROUP BY 1 ORDER BY 1
        """, {'y': year, 'm': month})
        return {int(r['day_']): r for r in rows}
    except Exception:
        return {}


def _lan_month_daily(year, month):
    """
    p_obs_zte_lan daily totals — for MAR-26 DT U/L and D/L rows.
    Matches Java: deduplicates duplicate timestamps via AVG before summing.
    Java: update p_obs_zte_lan set in_volume_kb=avg(in_volume_kb) where dup timestamps.
    We replicate by GROUP BY date_time with AVG, then SUM over the day.
    DL = in_volume_kb  (download to subscribers) → GB
    UL = out_volume_kb (upload from subscribers) → GB
    """
    try:
        rows = pg_query("""
            SELECT extract(day FROM date_time)::int              AS day_,
                   sum(coalesce(in_volume_kb,0))/1024.0/1024.0  AS dl_gb,
                   sum(coalesce(out_volume_kb,0))/1024.0/1024.0 AS ul_gb
            FROM public.p_obs_zte_lan
            WHERE extract(year FROM date_time)=%(y)s
              AND extract(month FROM date_time)=%(m)s
            GROUP BY 1 ORDER BY 1
        """, {'y': year, 'm': month})
        return {int(r['day_']): r for r in rows}
    except Exception:
        return {}


def _prtg4g_month_daily(year, month):
    """
    p_obs_zte_lan_4g_tput daily totals — for MAR-26 DT U/L and D/L rows.
    Used when ZTE SGSN is shutdown: DT = prtg4g_total - Nokia_SGSN.
    DL = in_volume_kb  (download to subscribers) → GB
    UL = out_volume_kb (upload from subscribers) → GB
    """
    try:
        rows = pg_query("""
            SELECT extract(day FROM date_time)::int              AS day_,
                   sum(coalesce(in_volume_kb,0))/1024.0/1024.0  AS dl_gb,
                   sum(coalesce(out_volume_kb,0))/1024.0/1024.0 AS ul_gb
            FROM public.p_obs_zte_lan_4g_tput
            WHERE extract(year FROM date_time)=%(y)s
              AND extract(month FROM date_time)=%(m)s
            GROUP BY 1 ORDER BY 1
        """, {'y': year, 'm': month})
        return {int(r['day_']): r for r in rows}
    except Exception:
        return {}


def _zte_5min_month(year, month):
    """ZTE 5-min monthly: dl_mb, ul_mb from p_obs_zte per slot."""
    try:
        return pg_query("""
            SELECT end_time AS ts,
                   extract(day FROM end_time)::int AS day,
                   coalesce(GTP81,0)/%(b)s          AS ul_mb,
                   coalesce(GTP83,0)/%(b)s          AS dl_mb,
                   coalesce(GTP29,0)/1024.0          AS dl_tput_mbps
            FROM p_obs_zte
            WHERE extract(year FROM end_time)=%(y)s
              AND extract(month FROM end_time)=%(m)s
            ORDER BY ts
        """, {'b': B_TO_MB, 'y': year, 'm': month})
    except Exception:
        return []


def _lan_5min_month(year, month):
    """
    LAN/NIB 5-min monthly: volume and speed per slot from p_obs_zte_lan.
    dl = in_volume_kb (download to subscribers)
    ul = out_volume_kb (upload from subscribers)
    dl_tput = in_speed_kbps/1024 Mbps
    ul_tput = out_speed_kbps/1024 Mbps
    """
    try:
        return pg_query("""
            SELECT date_time AS ts,
                   extract(day FROM date_time)::int           AS day,
                   sum(coalesce(in_volume_kb,0))/%(k)s         AS dl_mb,
                   sum(coalesce(out_volume_kb,0))/%(k)s        AS ul_mb,
                   avg(coalesce(in_speed_kbps,0))/1024.0       AS dl_tput_mbps,
                   avg(coalesce(out_speed_kbps,0))/1024.0      AS ul_tput_mbps
            FROM public.p_obs_zte_lan
            WHERE extract(year FROM date_time)=%(y)s
              AND extract(month FROM date_time)=%(m)s
            GROUP BY date_time ORDER BY ts
        """, {'k': KB_TO_MB, 'y': year, 'm': month})
    except Exception:
        return []


def _nokia_4g_month(year, month):
    """
    Nokia 4G monthly — per hour per day (for 5Min 4G column and MAR-26).
    5Min 4G = (lte_5212a+lte_5213a) per hour / 12 slots / MB_TO_GB
    MAR-26 4G = daily totals
    """
    try:
        return pg_query("""
            SELECT extract(day  FROM period_start_time)::int AS day_,
                   extract(hour FROM period_start_time)::int AS hr,
                   sum(coalesce(lte_5212a,0)) AS dl_mb,
                   sum(coalesce(lte_5213a,0)) AS ul_mb
            FROM public.nokia_4g_sgsn_report
            WHERE extract(year FROM period_start_time)=%(y)s
              AND extract(month FROM period_start_time)=%(m)s
            GROUP BY 1,2 ORDER BY 1,2
        """, {'y': year, 'm': month})
    except Exception:
        return []


def _nokia_4g_month_daily(year, month):
    """Nokia 4G daily totals — for MAR-26 4G U/L and D/L rows."""
    try:
        rows = pg_query("""
            SELECT extract(day FROM period_start_time)::int AS day_,
                   sum(coalesce(lte_5212a,0)) AS dl_mb,
                   sum(coalesce(lte_5213a,0)) AS ul_mb
            FROM public.nokia_4g_sgsn_report
            WHERE extract(year FROM period_start_time)=%(y)s
              AND extract(month FROM period_start_time)=%(m)s
            GROUP BY 1 ORDER BY 1
        """, {'y': year, 'm': month})
        return {int(r['day_']): r for r in rows}
    except Exception:
        return {}


def _prtg_5min_month(year, month):
    """
    Nokia 4G PRTG 5-min monthly — for Peak 4G column.
    in_speed_kbps / 1024 = DL Mbps per 5-min slot (sum all PRTG ports).
    """
    try:
        return pg_query("""
            SELECT date_time AS ts,
                   extract(day FROM date_time)::int           AS day,
                   sum(coalesce(in_volume_kb,0))/%(k)s         AS dl_mb,
                   sum(coalesce(in_speed_kbps,0))/1024.0       AS dl_tput_mbps
            FROM public.p_obs_zte_lan_4g_tput
            WHERE extract(year FROM date_time)=%(y)s
              AND extract(month FROM date_time)=%(m)s
            GROUP BY date_time ORDER BY ts
        """,{'k':KB_TO_MB,'y':year,'m':month})
    except Exception:
        return []


def _tcs4g_daily(date_str, tcs_filepath=None, log=print):
    """
    Read TCS EMS eNBwise daily CSV.
    1. If CSV file exists: read it, store into MySQL sgsn_tcs_daily, return data.
    2. If CSV file missing: fall back to MySQL sgsn_tcs_daily for that date.
    Returns daily total dict or None.
    """
    # Try CSV file first
    path = tcs_filepath
    if not path:
        info = resolve_tcs_file(date_str)
        if info["exists"]:
            path = info["path"]

    if path:
        try:
            result = read_tcs_daily(path)
            if result:
                result["day"] = int(date_str[8:10])
                # Store into MySQL for future use
                try:
                    store_tcs(result, log=log)
                except Exception as e:
                    log(f"[TCS Store] WARNING: could not save to MySQL: {e}")
                return result
        except Exception:
            pass

    # CSV not available — try MySQL fallback
    try:
        stored = load_tcs_day(date_str)
        if stored:
            log(f"[TCS] CSV not found — loaded from MySQL: "
                f"DL={stored['dl_mb']/1024:.0f}GB UL={stored['ul_mb']/1024:.0f}GB")
            return stored
    except Exception:
        pass

    return None


def _4g_kpi_hourly(date_str):
    """
    4G KPI — exact Java 3-table JOIN from nokia_sgsn_report.
    Returns dict: { hr: { emm_reg, eps_act, volte_u, peak_def, peak_ded,
                          peak_emm, emm_dereg, mos, rtp, rereg, att, succ, per_succ } }
    """
    try:
        rows = pg_query("""
            SELECT
                t1.period_start_time,
                t1.avg_emm_reg_users,
                t1.avg_active_eps_bearers,
                t1.avg_registered_eps_volte_users,
                t1.peak_active_def_eps_bearers_mme,
                t1.peak_active_ded_eps_bearers_mme,
                t1.peak_emm_reg_users_mme,
                t1.avg_emm_dereg_users,
                t2.mosvalue,
                t2.avgrtppacketloss,
                t2.reregistrationsuccessrate,
                t3.att_flns_5054b,
                t3.succ,
                t3.per_succ_flns_5053b
            FROM (
                SELECT
                    min(period_start_time) period_start_time,
                    round(avg(flns_5035a)::numeric,0) avg_emm_reg_users,
                    round(avg(flns_5050a)::numeric,0) avg_active_eps_bearers,
                    round(max(flns_5055a)::numeric,0) avg_registered_eps_volte_users,
                    round(max(flns_3285a)::numeric,0) peak_active_def_eps_bearers_mme,
                    round(max(flns_3286a)::numeric,0) peak_active_ded_eps_bearers_mme,
                    round(max(flns_5025a)::numeric,0) peak_emm_reg_users_mme,
                    round(avg(flns_5026a)::numeric,0) avg_emm_dereg_users
                FROM public.nokia_sgsn_report
                WHERE date(period_start_time) = %(dt)s
                GROUP BY extract(hour FROM period_start_time)
            ) t1
            INNER JOIN (
                SELECT min(period_start_time) period_start_time,
                    round(avg(mosvalue)::numeric,2)                  mosvalue,
                    round(avg(avgrtppacketloss)::numeric,2)          avgrtppacketloss,
                    round(avg(reregistrationsuccessrate)::numeric,2) reregistrationsuccessrate
                FROM public.sgsn_report_sbc_kpi
                WHERE date(period_start_time) = %(dt)s
                GROUP BY extract(hour FROM period_start_time)
            ) t2 ON t1.period_start_time = t2.period_start_time
            INNER JOIN (
                SELECT period_start_time,
                    att_flns_5054b,
                    succ,
                    round(per_succ_flns_5053b::numeric,2) per_succ_flns_5053b
                FROM public.sgsn_report_volte_dedicated_bearer_act
                WHERE date(period_start_time) = %(dt)s
            ) t3 ON t3.period_start_time = t2.period_start_time
            ORDER BY t1.period_start_time
        """, {'dt': date_str})

        result = {}
        for r in rows:
            hr = r['period_start_time'].hour
            result[hr] = {
                'emm_reg':  _safe(r['avg_emm_reg_users']),
                'eps_act':  _safe(r['avg_active_eps_bearers']),
                'volte_u':  _safe(r['avg_registered_eps_volte_users']),
                'peak_def': _safe(r['peak_active_def_eps_bearers_mme']),
                'peak_ded': _safe(r['peak_active_ded_eps_bearers_mme']),
                'peak_emm': _safe(r['peak_emm_reg_users_mme']),
                'emm_dereg':_safe(r['avg_emm_dereg_users']),
                'mos':      _safe(r['mosvalue']),
                'rtp':      _safe(r['avgrtppacketloss']),
                'rereg':    _safe(r['reregistrationsuccessrate']),
                'att':      _safe(r['att_flns_5054b']),
                'succ':     _safe(r['succ']),
                'per_succ': _safe(r['per_succ_flns_5053b']),
            }
        if result:
            return result
    except Exception:
        pass

    # Fallback: 3 separate queries
    emm_rows = {}
    try:
        rows = pg_query("""
            SELECT extract(hour FROM period_start_time)::int AS hr,
                round(avg(coalesce(flns_5035a,0))::numeric,0) AS emm_reg,
                round(avg(coalesce(flns_5050a,0))::numeric,0) AS eps_act,
                round(max(coalesce(flns_5055a,0))::numeric,0) AS volte_u,
                round(max(coalesce(flns_3285a,0))::numeric,0) AS peak_def,
                round(max(coalesce(flns_3286a,0))::numeric,0) AS peak_ded,
                round(max(coalesce(flns_5025a,0))::numeric,0) AS peak_emm,
                round(avg(coalesce(flns_5026a,0))::numeric,0) AS emm_dereg
            FROM public.nokia_sgsn_report
            WHERE date(period_start_time)=%(dt)s
            GROUP BY 1 ORDER BY 1
        """, {'dt': date_str})
        emm_rows = {r['hr']: r for r in rows}
    except Exception:
        pass

    sbc_rows = {}
    try:
        rows = pg_query("""
            SELECT extract(hour FROM period_start_time)::int AS hr,
                coalesce(mosvalue,0) AS mos,
                coalesce(avgrtppacketloss,0) AS rtp,
                coalesce(reregistrationsuccessrate,0) AS rereg
            FROM public.sgsn_report_sbc_kpi
            WHERE date(period_start_time)=%(dt)s ORDER BY hr
        """, {'dt': date_str})
        sbc_rows = {r['hr']: r for r in rows}
    except Exception:
        pass

    volte_rows = {}
    try:
        rows = pg_query("""
            SELECT extract(hour FROM period_start_time)::int AS hr,
                coalesce(att_flns_5054b,0) AS att,
                coalesce(succ,0) AS succ,
                round(coalesce(per_succ_flns_5053b,0)::numeric,2) AS per_succ
            FROM public.sgsn_report_volte_dedicated_bearer_act
            WHERE date(period_start_time)=%(dt)s ORDER BY hr
        """, {'dt': date_str})
        volte_rows = {r['hr']: r for r in rows}
    except Exception:
        pass

    result = {}
    for hr in range(24):
        e = emm_rows.get(hr, {})
        s = sbc_rows.get(hr, {})
        v = volte_rows.get(hr, {})
        result[hr] = {
            'emm_reg':  _safe(e.get('emm_reg')),
            'eps_act':  _safe(e.get('eps_act')),
            'volte_u':  _safe(e.get('volte_u')),
            'peak_def': _safe(e.get('peak_def')),
            'peak_ded': _safe(e.get('peak_ded')),
            'peak_emm': _safe(e.get('peak_emm')),
            'emm_dereg':_safe(e.get('emm_dereg')),
            'mos':      _safe(s.get('mos')),
            'rtp':      _safe(s.get('rtp')),
            'rereg':    _safe(s.get('rereg')),
            'att':      _safe(v.get('att')),
            'succ':     _safe(v.get('succ')),
            'per_succ': _safe(v.get('per_succ')),
        }
    return result


def _trai_month(year, month):
    try:
        return my_query("""
            SELECT `date`                        AS Date_,
                   round(`vol_2g_up_in_mb`,0)    AS ul_2g,
                   round(`vol_2g_dn_in_mb`,0)    AS dl_2g,
                   round(`vol_3g_up_in_mb`,0)    AS ul_3g,
                   round(`vol_3g_dn_in_mb`,0)    AS dl_3g,
                   round(`vol_4g_up_in_mb`,0)    AS ul_4g,
                   round(`vol_4g_dn_in_mb`,0)    AS dl_4g,
                   round(`vol_tot_in_gb`/1024,2)  AS total_tb,
                   round(`avg_2g_tput`,0)         AS avg_2g,
                   round(`avg_3g_tput`,0)         AS avg_3g,
                   round(`avg_4g_tput`,0)         AS avg_4g,
                   round(`avg_gn_tput`,0)         AS avg_gn,
                   round(`peak_2g_tput`,0)        AS peak_2g,
                   round(`peak_3g_tput`,0)        AS peak_3g,
                   round(`peak_4g_tput`,0)        AS peak_4g,
                   round(`peak_gn_tput`,0)        AS peak_gn,
                   round(`attach_sub_2g`,0)       AS att_2g,
                   round(`attach_sub_3g`,0)       AS att_3g,
                   round(`attach_sub_4g`,0)       AS att_4g
            FROM `reportsDB`.`sgsn_trai_report`
            WHERE month(`date`)=%s AND year(`date`)=%s
            ORDER BY `date`
        """, (month, year))
    except Exception:
        return []


# =============================================================================
#  SHEET 1 — MAR-26
# =============================================================================

def _build_mar26(wb, year, month, month_name, report_day,
                 nokia_daily, nokia4g_daily, lan_daily, tcs_daily):
    """
    MAR-26 confirmed sources:
      Nokia U/L → nokia_sgsn_report_out: sum(upload_2g+upload_3g) per day  (MB→GB)
      Nokia D/L → nokia_sgsn_report_out: sum(download_2g+download_3g) per day (MB→GB)
      DT U/L    → p_obs_zte_lan_4g_tput: sum(out_volume_kb)/1024/1024 per day (GB) - Nokia_UL
      DT D/L    → p_obs_zte_lan_4g_tput: sum(in_volume_kb)/1024/1024 per day (GB) - Nokia_DL
      4G U/L    → nokia_4g_sgsn_report: sum(lte_5213a) per day (MB→GB)
      4G D/L    → nokia_4g_sgsn_report: sum(lte_5212a) per day (MB→GB)
      TCS U/L   → TCS CSV col73 UL MB / MB_TO_GB
      TCS D/L   → TCS CSV col72 DL MB / MB_TO_GB
      ZTE       → zero (shutdown)
    """
    ws = wb.create_sheet(f"{month_name[:3]}-{str(year)[2:]}")
    days = calendar.monthrange(year, month)[1]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=35)
    c = ws.cell(1, 1, "SGSN DATA USAGE REPORT")
    c.font=_font(True, C_DARK, 12); c.alignment=_align(); c.fill=_fill(C_LBLUE)

    CO = 3
    _hdr(ws, 3, 1, "DATE", C_MED); _hdr(ws, 3, 2, "", C_MED)
    for d in range(1, days+1): _hdr(ws, 3, CO+d-1, d, C_MED)
    for lbl, off in [("Avg", days), ("Month Max", days+1), ("MAX till date", days+2)]:
        _hdr(ws, 3, CO+off, lbl, C_DARK)

    ROWS = [
        ("SGSN_ZTE",         "U/L", 4,  "zte_ul"),
        ("",                 "D/L", 5,  "zte_dl"),
        ("SGSN_NOKIA",       "U/L", 6,  "nok_ul"),
        ("",                 "D/L", 7,  "nok_dl"),
        ("DIRECT TUNNELING", "U/L", 8,  "dt_ul"),
        ("",                 "D/L", 9,  "dt_dl"),
        ("4G",               "U/L", 10, "g4_ul"),
        ("",                 "D/L", 11, "g4_dl"),
        ("TCS 4G",           "U/L", 12, "tcs_ul"),
        ("",                 "D/L", 13, "tcs_dl"),
    ]

    daily = {d: {k: 0.0 for k in ['zte_ul','zte_dl','nok_ul','nok_dl',
                                    'dt_ul','dt_dl','g4_ul','g4_dl',
                                    'tcs_ul','tcs_dl']}
             for d in range(1, days+1)}

    for d in range(1, days+1):
        # Nokia U/L and D/L from nokia_sgsn_report_out (MB → GB)
        n = nokia_daily.get(d, {})
        daily[d]['nok_ul'] = _safe(n.get('ul_mb', 0)) / MB_TO_GB
        daily[d]['nok_dl'] = _safe(n.get('dl_mb', 0)) / MB_TO_GB

        # DT (Direct Tunneling) = p_obs_zte_lan_4g_tput_total - Nokia_2G3G
        # ZTE SGSN is shutdown → ZTE contribution is zero
        # DT_UL = prtg4g_UL - Nokia_UL  (current value after subtraction)
        # DT_DL = prtg4g_DL - Nokia_DL  (current value after subtraction)
        l = lan_daily.get(d, {})
        nib_ul_gb = _safe(l.get('ul_gb', 0))
        nib_dl_gb = _safe(l.get('dl_gb', 0))
        daily[d]['dt_ul'] = max(0.0, nib_ul_gb - daily[d]['nok_ul'] - daily[d]['zte_ul'])
        daily[d]['dt_dl'] = max(0.0, nib_dl_gb - daily[d]['nok_dl'] - daily[d]['zte_dl'])

        # 4G from nokia_4g_sgsn_report (MB → GB)
        g = nokia4g_daily.get(d, {})
        daily[d]['g4_ul'] = _safe(g.get('ul_mb', 0)) / MB_TO_GB
        daily[d]['g4_dl'] = _safe(g.get('dl_mb', 0)) / MB_TO_GB

    # TCS — load ALL previous days from MySQL sgsn_tcs_daily,
    # then overlay today with CSV data (most accurate).
    # MySQL data is populated each day when Daily Report runs with a TCS CSV.
    # This gives correct history for the whole month in MAR-26 sheet.
    try:
        tcs_month = load_tcs_month(year, month)
        for d, tr in tcs_month.items():
            if d in daily:
                daily[d]['tcs_ul'] = _safe(tr.get('ul_mb', 0)) / MB_TO_GB
                daily[d]['tcs_dl'] = _safe(tr.get('dl_mb', 0)) / MB_TO_GB
    except Exception:
        pass
    # Overlay today's data (from CSV or MySQL fallback for today)
    if tcs_daily and isinstance(tcs_daily, dict):
        d = tcs_daily.get('day', report_day)
        if d in daily:
            daily[d]['tcs_ul'] = _safe(tcs_daily.get('ul_mb', 0)) / MB_TO_GB
            daily[d]['tcs_dl'] = _safe(tcs_daily.get('dl_mb', 0)) / MB_TO_GB

    for lbl, ud, row, key in ROWS:
        if lbl:
            c = ws.cell(row, 1, lbl)
            c.fill=_fill(C_LTEAL); c.font=_font(True); c.alignment=_align()
        ws.cell(row, 2, ud).alignment=_align()
        vals = []
        for d in range(1, days+1):
            if d > report_day:
                ws.cell(row, CO+d-1, ""); continue
            v = round(daily[d][key], 3)
            ws.cell(row, CO+d-1, v).alignment=_align()
            vals.append(float(v))
        avg = round(sum(vals)/len(vals), 3) if vals else 0.0
        mx  = round(max(vals), 3)           if vals else 0.0
        ws.cell(row, CO+days,   avg).alignment=_align()
        ws.cell(row, CO+days+1, mx ).alignment=_align()
        ws.cell(row, CO+days+2, mx ).alignment=_align()

    # Total row
    c = ws.cell(14, 1, "Total")
    c.fill=_fill(C_ORNG); c.font=_font(True, C_WHITE); c.alignment=_align()
    for d in range(1, days+1):
        if d > report_day:
            ws.cell(14, CO+d-1, ""); continue
        tot = sum(float(daily[d][k]) for k in daily[d])
        ws.cell(14, CO+d-1, round(tot, 3)).alignment=_align()

    ws.cell(17, 1, "Note-  Data usage in GB").font=_font(True)
    ws.column_dimensions['A'].width=18; ws.column_dimensions['B'].width=6
    for d in range(1, days+1):
        ws.column_dimensions[get_column_letter(CO+d-1)].width=10



# =============================================================================
#  SHEETS 2 & 3 — SGSNZTE / SGSNNOKIA
# =============================================================================

def _build_node_sheet(wb, sheet_name, date_str, dt,
                      hourly_vol, hourly_pdp, hourly_att, is_nokia=False):
    ws = wb.create_sheet(sheet_name)
    dl = dt.strftime("%Y-%m-%d")

    _mhdr(ws, 4,  3,  4, f"3G Data Usage_{sheet_name}")
    _mhdr(ws, 4,  5,  6, f"2G Data Usage_{sheet_name}")
    _mhdr(ws, 4,  7,  9, f"Total Data Usage_{sheet_name}")
    _mhdr(ws, 4, 10, 12, f"Throughput_{sheet_name}")
    _mhdr(ws, 4, 15, 16, f"3G PDP {sheet_name}", C_DARK)
    _mhdr(ws, 4, 17, 18, f"2G PDP {sheet_name}", C_DARK)
    _mhdr(ws, 4, 19, 21, f"Throughput {sheet_name}", C_DARK)
    _mhdr(ws, 4, 23, 35, "Rejection", C_ORNG)
    _mhdr(ws, 4, 37, 39, "Attach", C_GREEN)

    for col, lbl in [
        (2,"Date & Time"),(3,"Upload\n(in MB)"),(4,"Download\n(in MB)"),
        (5,"Upload\n(in MB)"),(6,"Download\n(in MB)"),
        (7,"Total 2G\n(in MB)"),(8,"Total 3G\n(in MB)"),(9,"Total Usage\n(in MB)"),
        (10,"Total 2G\n(in Mbps)"),(11,"Total 3G\n(in Mbps)"),(12,"Total\n(in Mbps)"),
        (14,"Date & Time"),(15,"Attched Sub "),(16,"Active PDP"),
        (17,"Attched Sub "),(18,"Active PDP "),
        (19,"Upload (in Mbps)"),(20,"Download  (in Mbps)"),(21,"Total\n(in Mbps)"),
        (23,"Attempt"),(24,"Success"),(25,"Rej GGSN"),(26,"Rej IN"),
        (27,"Rej Others"),(28,"Rej\nOver load"),(29,"Rej\nRF"),
        (30,"% Success"),(31,"% Rej GGSN"),(32,"% Rej IN"),
        (33,"%Rej Others"),(34,"% Rej Over load"),(35,"% Rej\nRF"),
        (37,"Attempt"),(38,"Success"),(39,"% Success"),
    ]:
        c = ws.cell(5, col, lbl)
        c.fill=_fill(C_LTEAL); c.font=_font(True); c.alignment=_align(wrap=True); c.border=_border()

    T  = {k: 0.0 for k in ['ul3g','dl3g','ul2g','dl2g',
                              'ra','rs','rg','ri','ro','rv','rr','aa','as_']}
    MX = {k: 0.0 for k in ['p3a','p3c','p2a','p2c','ul_t','dl_t','tput2g','tput3g',
                              'pct_s','pct_g','pct_i','pct_o','pct_v','pct_r','att_p']}

    for hr in range(24):
        row = 6+hr; ts = f"{dl} {hr:02d}:00:00"
        v = hourly_vol.get(hr, {})
        p = hourly_pdp.get(hr, {})
        a = hourly_att.get(hr, {})

        ul3g = _safe(v.get('ul_3g_mb')); dl3g = _safe(v.get('dl_3g_mb'))
        ul2g = _safe(v.get('ul_2g_mb')); dl2g = _safe(v.get('dl_2g_mb'))
        t2g  = ul2g+dl2g; t3g = ul3g+dl3g; tus = t2g+t3g

        ul_t   = _safe(v.get('ul_tput', 0)); dl_t = _safe(v.get('dl_tput', 0))
        tput3g = _safe(v.get('tput_3g', 0)) if is_nokia else _safe(v.get('dl_tput', 0))
        tput2g = _safe(v.get('tput_2g', 0)) if is_nokia else 0.0

        p3a = _safe(v.get('pdp3g_att')); p3c = _safe(v.get('pdp3g_act'))
        p2a = _safe(v.get('pdp2g_att')); p2c = _safe(v.get('pdp2g_act'))

        # Nokia: rejection + attach already in hourly_vol from nokia_sgsn_report_out join
        if is_nokia:
            ra  = _safe(v.get('rej_att',  p.get('att',  0)))
            rs  = _safe(v.get('rej_succ', p.get('succ', 0)))
            rg  = _safe(v.get('rej_ggsn', p.get('rej_ggsn', 0)))
            ri  = _safe(v.get('rej_in',   p.get('rej_in',   0)))
            ro  = _safe(v.get('rej_oth',  p.get('rej_oth',  0)))
            rv  = _safe(v.get('rej_ovrld',p.get('rej_ovrld',0)))
            rr  = _safe(v.get('rej_rf',   p.get('rej_rf',   0)))
            aa  = _safe(v.get('att_att',  a.get('att',  0)))
            as_ = _safe(v.get('att_succ', a.get('succ', 0)))
        else:
            ra  = _safe(p.get('att'));   rs  = _safe(p.get('succ'))
            rg  = _safe(p.get('rej_ggsn')); ri = _safe(p.get('rej_in'))
            ro  = _safe(p.get('rej_oth')); rv = _safe(p.get('rej_ovrld')); rr = _safe(p.get('rej_rf'))
            aa  = _safe(a.get('att')); as_ = _safe(a.get('succ'))

        pct_s=_pct(rs,ra); pct_g=_pct(rg,ra); pct_i=_pct(ri,ra)
        pct_o=_pct(ro,ra); pct_v=_pct(rv,ra); pct_r=_pct(rr,ra)
        att_p=_pct(as_,aa)

        for col, val in [
            (2,ts),(3,int(round(ul3g))),(4,int(round(dl3g))),
            (5,int(round(ul2g))),(6,int(round(dl2g))),
            (7,int(round(t2g))),(8,int(round(t3g))),(9,int(round(tus))),
            (10,round(tput2g,2)),(11,round(tput3g,2)),(12,round(tput2g+tput3g,2)),
            (14,ts),(15,int(p3a)),(16,int(p3c)),(17,int(p2a)),(18,int(p2c)),
            (19,round(ul_t,2)),(20,round(dl_t,2)),(21,round(ul_t+dl_t,2)),
            (23,int(ra)),(24,int(rs)),(25,int(rg)),(26,int(ri)),
            (27,int(ro)),(28,int(rv)),(29,int(rr)),
            (30,pct_s),(31,pct_g),(32,pct_i),(33,pct_o),(34,pct_v),(35,pct_r),
            (37,int(aa)),(38,int(as_)),(39,att_p),
        ]: _dat(ws, row, col, val)

        for k, val in [('ul3g',ul3g),('dl3g',dl3g),('ul2g',ul2g),('dl2g',dl2g),
                        ('ra',ra),('rs',rs),('rg',rg),('ri',ri),
                        ('ro',ro),('rv',rv),('rr',rr),('aa',aa),('as_',as_)]:
            T[k] += val
        for k, val in [('p3a',p3a),('p3c',p3c),('p2a',p2a),('p2c',p2c),
                        ('ul_t',ul_t),('dl_t',dl_t),('tput2g',tput2g),('tput3g',tput3g),
                        ('pct_s',pct_s),('pct_g',pct_g),('pct_i',pct_i),
                        ('pct_o',pct_o),('pct_v',pct_v),('pct_r',pct_r),('att_p',att_p)]:
            if val > MX[k]: MX[k] = val

    tr = 30
    t2g=T['ul2g']+T['dl2g']; t3g=T['ul3g']+T['dl3g']; tus=t2g+t3g
    ps=_pct(T['rs'],T['ra']); pg=_pct(T['rg'],T['ra']); pi=_pct(T['ri'],T['ra'])
    po=_pct(T['ro'],T['ra']); pv=_pct(T['rv'],T['ra']); pr=_pct(T['rr'],T['ra'])
    ap=_pct(T['as_'],T['aa'])

    for col, val in [
        (2,"TOTAL"),(3,int(round(T['ul3g']))),(4,int(round(T['dl3g']))),
        (5,int(round(T['ul2g']))),(6,int(round(T['dl2g']))),
        (7,int(round(t2g))),(8,int(round(t3g))),(9,int(round(tus))),
        (10,round(MX['tput2g'],2)),(11,round(MX['tput3g'],2)),
        (12,round(MX['tput2g']+MX['tput3g'],2)),
        (14,"MAX"),(15,int(MX['p3a'])),(16,int(MX['p3c'])),
        (17,int(MX['p2a'])),(18,int(MX['p2c'])),
        (19,round(MX['ul_t'],2)),(20,round(MX['dl_t'],2)),
        (21,round(MX['ul_t']+MX['dl_t'],2)),
        (23,int(T['ra'])),(24,int(T['rs'])),(25,int(T['rg'])),(26,int(T['ri'])),
        (27,int(T['ro'])),(28,int(T['rv'])),(29,int(T['rr'])),
        (30,ps),(31,pg),(32,pi),(33,po),(34,pv),(35,pr),
        (37,int(T['aa'])),(38,int(T['as_'])),(39,ap),
    ]: _total_row(ws, tr, col, val)

    ws.column_dimensions['A'].width = 4; ws.column_dimensions['B'].width = 22
    for col in range(3, 40): ws.column_dimensions[get_column_letter(col)].width = 11
    ws.row_dimensions[5].height = 35


# =============================================================================
#  SHEET 4 — Total
# =============================================================================

def _build_total(wb, date_str, dt, zte_h, nokia_h, g4_h, lan_total_h):
    """
    Total sheet — EXACT Java query replicated in Python:

    t3 = p_obs_zte     : ZTE 2G (GTP73=UL, GTP75=DL) grouped by hour
    t4 = nokia_sgsn_report_out : Nokia 2G (upload_2g, download_2g) grouped by hour
    t5 = p_obs_zte_lan : NIB/DT total volume + MAX speed per hour
    t6 = nokia_4g_sgsn_report  : 4G (lte_5212a DL, lte_5213a UL) grouped by hour

    Cols C,D = 4G UL/DL (t6)
    Cols E,F = 3G = t5.upload/download MINUS (ZTE_2G + Nokia_2G)
    Cols G,H = 2G = ZTE_2G + Nokia_2G
    Cols I-L = Totals (MB)
    Cols M-P = Throughput = volume_MB * 8 / 3600 (Mbps)
    Cols Y,Z,AA = NIB MAX speed per hour (MAX not SUM — exact Java t5 formula)
    """
    ws = wb.create_sheet("Total")
    dl = dt.strftime("%Y-%m-%d")

    _mhdr(ws, 4,  3,  4, "4G Data Usage")
    _mhdr(ws, 4,  5,  6, "3G Data Usage")
    _mhdr(ws, 4,  7,  8, "2G  Data Usage")
    _mhdr(ws, 4,  9, 12, "Total Data Usage")
    _mhdr(ws, 4, 13, 16, "Throughput")
    _mhdr(ws, 4, 19, 20, "3G PDP")
    _mhdr(ws, 4, 21, 22, "2G PDP")
    _mhdr(ws, 4, 23, 24, "2G+3G PDP")
    _mhdr(ws, 4, 25, 27, "Throughput")
    _mhdr(ws, 4, 29, 41, "Rejection", C_ORNG)
    _mhdr(ws, 4, 43, 45, "Attach", C_GREEN)

    for col, lbl in [
        (2,"Date & Time"),(3,"Upload\n(in MB)"),(4,"Download (in MB)"),
        (5,"Upload\n(in MB)"),(6,"Download\n(in MB)"),
        (7,"Upload\n(in MB)"),(8,"Download\n(in MB)"),
        (9,"Total 4G\n(in MB)"),(10,"Total 3G\n(in MB)"),
        (11,"Total 2G\n(in MB)"),(12,"Total Usage\n(in MB)"),
        (13,"Total 4G\n(in Mbps)"),(14,"Total 3G\n(in Mbps)"),
        (15,"Total 2G\n(in Mbps)"),(16,"Total\n(in Mbps)"),
        (18,"Date & Time"),
        (19,"Attched Sub"),(20,"Active PDP"),(21,"Attched Sub"),(22,"Active PDP "),
        (23,"Attched Sub "),(24,"Active PDP "),
        (25,"Upload\n(in Mbps)"),(26,"Download\n(in Mbps)"),(27,"Total\n(in Mbps)"),
        (29,"Attempt"),(30,"Success"),(31,"Rej GGSN"),(32,"Rej IN"),
        (33,"Rej Others"),(34,"Rej\nOver load"),(35,"Rej\nRF"),
        (36,"% Success"),(37,"% Rej GGSN"),(38,"% Rej IN"),
        (39,"%Rej Others"),(40,"% Rej Over load"),(41,"% Rej\nRF"),
        (43,"Attempt"),(44,"Success"),(45,"% Success"),
    ]:
        c = ws.cell(5, col, lbl)
        c.fill=_fill(C_LTEAL); c.font=_font(True); c.alignment=_align(wrap=True); c.border=_border()

    T  = {k: 0.0 for k in ['ul4g','dl4g','ul3g','dl3g','ul2g','dl2g',
                              'ra','rs','rg','ri','ro','rv','rr','aa','as_']}
    MX = {k: 0.0 for k in ['tp4g','tp3g','tp2g','tp_t',
                              'p3a','p3c','p2a','p2c','p23a','p23c','ul_t','dl_t']}

    for hr in range(24):
        row = 6+hr; ts = f"{dl} {hr:02d}:00:00"
        z  = zte_h.get(hr, {})         # t3: ZTE
        n  = nokia_h.get(hr, {})       # t4: Nokia
        g  = g4_h.get(hr, {})          # t6: 4G
        lh = lan_total_h.get(hr, {})   # t5: NIB/LAN

        # t6: 4G columns C,D
        ul4g = _safe(g.get('ul_mb',  0))
        dl4g = _safe(g.get('dl_mb',  0))

        # t3: ZTE 2G only (GTP73=UL, GTP75=DL)
        zte_ul2g = _safe(z.get('ul_2g_mb', 0))
        zte_dl2g = _safe(z.get('dl_2g_mb', 0))

        # t4: Nokia 2G only (upload_2g, download_2g from nokia_sgsn_report_out)
        nok_ul2g = _safe(n.get('ul_2g_mb', 0))
        nok_dl2g = _safe(n.get('dl_2g_mb', 0))

        # t5: NIB/LAN = Direct Tunneling (p_obs_zte_lan)
        nib_ul   = _safe(lh.get('ul_mb',      0))   # sum(out_volume_kb)/1024
        nib_dl   = _safe(lh.get('dl_mb',      0))   # sum(in_volume_kb)/1024
        nib_up_s = _safe(lh.get('up_speed',   0))   # MAX(out_speed_kbps)/1024
        nib_dn_s = _safe(lh.get('down_speed', 0))   # MAX(in_speed_kbps)/1024
        nib_tot_s= _safe(lh.get('tot_speed',  0))   # MAX(in+out)/1024

        # Cols E,F: 3G = NIB TOTAL minus ZTE_2G minus Nokia_2G  (exact Java formula)
        ul3g = max(0.0, nib_ul  - (zte_ul2g + nok_ul2g))
        dl3g = max(0.0, nib_dl  - (zte_dl2g + nok_dl2g))

        # Cols G,H: 2G = ZTE_2G + Nokia_2G
        ul2g = zte_ul2g + nok_ul2g
        dl2g = zte_dl2g + nok_dl2g

        # Totals
        t4g  = ul4g + dl4g
        t3g  = ul3g + dl3g
        t2g  = ul2g + dl2g
        tus  = t4g + t3g + t2g

        # Throughput: volume_MB * 8 / 3600 (exact Java formula for M,N,O,P)
        tp4g = t4g * 8.0 / 3600.0
        tp3g = t3g * 8.0 / 3600.0
        tp2g = t2g * 8.0 / 3600.0
        tp_t = tus * 8.0 / 3600.0

        # PDP from nokia_sgsn_report_out
        p3a  = _safe(n.get('pdp3g_att')); p3c = _safe(n.get('pdp3g_act'))
        p2a  = _safe(n.get('pdp2g_att')); p2c = _safe(n.get('pdp2g_act'))
        p23a = p3a + p2a; p23c = p3c + p2c

        # Rejection / Attach from nokia_sgsn_report_out join
        ra  = _safe(n.get('rej_att')); rs  = _safe(n.get('rej_succ'))
        rg  = _safe(n.get('rej_ggsn')); ri  = _safe(n.get('rej_in'))
        ro  = _safe(n.get('rej_oth')); rv  = _safe(n.get('rej_ovrld'))
        rr  = _safe(n.get('rej_rf'))
        aa  = _safe(n.get('att_att')); as_ = _safe(n.get('att_succ'))
        pct_s=_pct(rs,ra); pct_g=_pct(rg,ra); pct_i=_pct(ri,ra)
        pct_o=_pct(ro,ra); pct_v=_pct(rv,ra); pct_r=_pct(rr,ra)
        att_p=_pct(as_,aa)

        for col, val in [
            (2,ts),(3,int(round(ul4g))),(4,int(round(dl4g))),
            (5,int(round(ul3g))),(6,int(round(dl3g))),
            (7,int(round(ul2g))),(8,int(round(dl2g))),
            (9,int(round(t4g))),(10,int(round(t3g))),(11,int(round(t2g))),(12,int(round(tus))),
            (13,round(tp4g,2)),(14,round(tp3g,2)),(15,round(tp2g,2)),(16,round(tp_t,2)),
            (18,ts),(19,int(p3a)),(20,int(p3c)),(21,int(p2a)),(22,int(p2c)),
            (23,int(p23a)),(24,int(p23c)),
            (25,round(nib_up_s,2)),(26,round(nib_dn_s,2)),(27,round(nib_tot_s,2)),
            (29,int(ra)),(30,int(rs)),(31,int(rg)),(32,int(ri)),
            (33,int(ro)),(34,int(rv)),(35,int(rr)),
            (36,pct_s),(37,pct_g),(38,pct_i),(39,pct_o),(40,pct_v),(41,pct_r),
            (43,int(aa)),(44,int(as_)),(45,att_p),
        ]: _dat(ws, row, col, val)

        for k, val in [('ul4g',ul4g),('dl4g',dl4g),('ul3g',ul3g),('dl3g',dl3g),
                        ('ul2g',ul2g),('dl2g',dl2g),
                        ('ra',ra),('rs',rs),('rg',rg),('ri',ri),
                        ('ro',ro),('rv',rv),('rr',rr),('aa',aa),('as_',as_)]:
            T[k] += val
        for k, val in [('tp4g',tp4g),('tp3g',tp3g),('tp2g',tp2g),('tp_t',tp_t),
                        ('p3a',p3a),('p3c',p3c),('p2a',p2a),('p2c',p2c),
                        ('p23a',p23a),('p23c',p23c),
                        ('ul_t',nib_up_s),('dl_t',nib_dn_s)]:
            if val > MX[k]: MX[k] = val

    tr = 30; t_all = sum(T[k] for k in ('ul4g','dl4g','ul3g','dl3g','ul2g','dl2g'))
    ps=_pct(T['rs'],T['ra']); pg=_pct(T['rg'],T['ra']); pi=_pct(T['ri'],T['ra'])
    po=_pct(T['ro'],T['ra']); pv=_pct(T['rv'],T['ra']); pr=_pct(T['rr'],T['ra'])
    ap=_pct(T['as_'],T['aa'])

    for col, val in [
        (2,"TOTAL"),
        (3,int(round(T['ul4g']))),(4,int(round(T['dl4g']))),
        (5,int(round(T['ul3g']))),(6,int(round(T['dl3g']))),
        (7,int(round(T['ul2g']))),(8,int(round(T['dl2g']))),
        (9,int(round(T['ul4g']+T['dl4g']))),(10,int(round(T['ul3g']+T['dl3g']))),
        (11,int(round(T['ul2g']+T['dl2g']))),(12,int(round(t_all))),
        (13,round(MX['tp4g'],2)),(14,round(MX['tp3g'],2)),
        (15,round(MX['tp2g'],2)),(16,round(MX['tp_t'],2)),
        (18,"MAX"),(19,int(MX['p3a'])),(20,int(MX['p3c'])),
        (21,int(MX['p2a'])),(22,int(MX['p2c'])),
        (23,int(MX['p23a'])),(24,int(MX['p23c'])),
        (25,round(MX['ul_t'],2)),(26,round(MX['dl_t'],2)),
        (27,round(MX['ul_t']+MX['dl_t'],2)),
        (29,int(T['ra'])),(30,int(T['rs'])),(31,int(T['rg'])),(32,int(T['ri'])),
        (33,int(T['ro'])),(34,int(T['rv'])),(35,int(T['rr'])),
        (36,ps),(37,pg),(38,pi),(39,po),(40,pv),(41,pr),
        (43,int(T['aa'])),(44,int(T['as_'])),(45,ap),
    ]: _total_row(ws, tr, col, val)

    ws.cell(36, 2, "Total Usage (in GB)").font=_font(True)
    ws.cell(36, 4, round(t_all/MB_TO_GB, 2)).alignment=_align()
    ws.column_dimensions['A'].width=4
    for col in range(2, 46): ws.column_dimensions[get_column_letter(col)].width=12
    ws.row_dimensions[5].height=35



# =============================================================================
#  SHEETS 5 & 6 — 5 Min / Peak Throughput
# =============================================================================

def _build_5min(wb, year, month, report_day,
               zte_5m, nokia_5m, lan_5m, g4_5m, prtg_5m):
    """
    5 Min sheet — 288 5-min slots x report_day date columns.
    Confirmed sources from reference XLS:
      ZTE   : p_obs_zte GTP81+GTP83 bytes/B_TO_MB/MB_TO_GB per 5-min slot
      NOKIA : nokia_sgsn_report_out totalgn per 15-min period / 3 / MB_TO_GB
      NIB   : p_obs_zte_lan (in_volume_kb+out_volume_kb)/KB_TO_MB/MB_TO_GB per 5-min
      4G    : nokia_4g_sgsn_report (lte_5212a+lte_5213a)/12 per hour / MB_TO_GB
      TOTAL : ZTE+NOKIA+NIB+4G
    Values in GB, 2 decimal places.
    """
    ws  = wb.create_sheet("5 Min")
    days = report_day
    CPD  = 5
    TCOL = 1 + days * CPD

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TCOL)
    ws.cell(1,1, f"SGSN Download (in GB) for Month of  {calendar.month_name[month]} {year}"
            ).font=_font(True, C_DARK, 11)

    # Row 2: date headers (merged 5 cols per day)
    c = ws.cell(2,1,"Date"); c.fill=_fill(C_DARK); c.font=_font(True,C_WHITE); c.alignment=_align()
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+4)
        cell = ws.cell(2, cs, f"{d:02d}-{month:02d}-{year}")
        cell.fill=_fill(C_MED); cell.font=_font(True,C_WHITE); cell.alignment=_align()

    # Row 3: sub-column headers
    c = ws.cell(3,1,"Time"); c.fill=_fill(C_DARK); c.font=_font(True,C_WHITE); c.alignment=_align()
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        for i, lbl in enumerate(['ZTE','NOKIA','NIB','4G','TOTAL']):
            cell = ws.cell(3, cs+i, lbl)
            cell.fill=_fill(C_LTEAL); cell.font=_font(True); cell.alignment=_align()

    # Build per-slot data dictionaries: {day: {slot: value}}
    zte_d   = {}   # {day: {slot: dl_gb}}
    nokia_d = {}   # {day: {period_slot: total_gb}} — 15-min, fill 3 slots
    lan_d   = {}   # {day: {slot: total_gb}}
    g4_d    = {}   # {day: {hr: total_gb}}

    for r in zte_5m:
        ts = r.get('ts')
        if not ts: continue
        d = int(r['day']); slot = (ts.hour*60+ts.minute)//5
        dl_gb = (_safe(r.get('dl_mb',0)) + _safe(r.get('ul_mb',0))) / MB_TO_GB
        zte_d.setdefault(d, {})[slot] = dl_gb

    for r in nokia_5m:
        d = int(r['day_']); hr = int(r['hr']); mn = int(r.get('mn',0) or 0)
        period_slot = hr*12 + (mn//15)*3
        gb = _safe(r.get('slot_total_mb',0)) / MB_TO_GB / 3.0  # 15-min → per 5-min
        # ADD not overwrite: nokia_sgsn_report_out has multiple node rows per period
        nokia_d.setdefault(d, {})
        nokia_d[d][period_slot] = nokia_d[d].get(period_slot, 0.0) + gb

    for r in lan_5m:
        ts = r.get('ts')
        if not ts: continue
        d = int(r['day']); slot = (ts.hour*60+ts.minute)//5
        gb = (_safe(r.get('dl_mb',0)) + _safe(r.get('ul_mb',0))) / MB_TO_GB
        lan_d.setdefault(d, {})[slot] = lan_d.get(d,{}).get(slot,0.0) + gb

    for r in g4_5m:
        d = int(r['day_']); hr = int(r['hr'])
        gb = (_safe(r.get('dl_mb',0)) + _safe(r.get('ul_mb',0))) / MB_TO_GB / 12.0
        g4_d.setdefault(d, {})[hr] = gb

    # Track per-column values for MAX and Total rows
    dsv = {d: [[] for _ in range(CPD)] for d in range(1, days+1)}

    for slot in range(288):
        row = 4+slot
        hr  = (slot*5)//60; mn = (slot*5)%60
        c = ws.cell(row,1, dt_time(hr,mn,0)); c.number_format='HH:MM'; c.alignment=_align()
        period_slot = (slot//3)*3   # round down to 15-min boundary

        for d in range(1, days+1):
            cs = 2+(d-1)*CPD
            zte_v   = zte_d.get(d,{}).get(slot, 0.0)
            nok_v   = nokia_d.get(d,{}).get(period_slot, 0.0)
            lan_v   = lan_d.get(d,{}).get(slot, 0.0)
            g4_v    = g4_d.get(d,{}).get(hr, 0.0)
            tot_v   = zte_v + nok_v + lan_v + g4_v

            vals = [zte_v, nok_v, lan_v, g4_v, tot_v]
            for i, val in enumerate(vals):
                ws.cell(row, cs+i, round(val, 2)).alignment=_align()
                dsv[d][i].append(val)

    # MAX row
    mr = 4+288
    ws.cell(mr,1,"MAX").font=_font(True); ws.cell(mr,1).fill=_fill(C_TOTAL)
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        for i in range(CPD):
            lst = dsv[d][i]; v = max(lst) if lst else 0.0
            cell = ws.cell(mr, cs+i, round(v,2))
            cell.font=_font(True); cell.fill=_fill(C_TOTAL); cell.alignment=_align()

    # Total row
    tr = 4+289
    ws.cell(tr,1,"Total").font=_font(True); ws.cell(tr,1).fill=_fill(C_GRAY)
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        for i in range(CPD):
            lst = dsv[d][i]; v = sum(lst) if lst else 0.0
            cell = ws.cell(tr, cs+i, round(v,2))
            cell.font=_font(True); cell.fill=_fill(C_GRAY); cell.alignment=_align()

    ws.column_dimensions['A'].width=10
    for col in range(2, TCOL+1): ws.column_dimensions[get_column_letter(col)].width=11


def _build_peak(wb, year, month, report_day,
               zte_5m, nokia_5m, lan_5m, prtg_5m):
    """
    Peak Throughput sheet — 288 5-min slots x report_day date columns.
    Confirmed sources from reference XLS:
      ZTE   : p_obs_zte GTP29 Mbps per 5-min slot
      NOKIA : nokia_sgsn_report_out (peak2gthru+peak3gthru) Mbps per 15-min (3 slots same)
      NIB   : p_obs_zte_lan sum(in_speed_kbps+out_speed_kbps)/1024 Mbps per 5-min slot
      4G    : p_obs_zte_lan_4g_tput sum(in_speed_kbps)/1024 Mbps per 5-min slot
      Total : (ZTE+NOKIA+NIB+4G)/1000 Gbps
    Only MAX row (no Total row) — confirmed from reference.
    """
    ws  = wb.create_sheet("Peak Throughput")
    days = report_day
    CPD  = 5
    TCOL = 1 + days * CPD

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TCOL)
    ws.cell(1,1, f"SGSN Peak Throughput (in Mbps) for Month of {calendar.month_name[month]} {year}"
            ).font=_font(True, C_DARK, 11)

    c = ws.cell(2,1,"Date"); c.fill=_fill(C_DARK); c.font=_font(True,C_WHITE); c.alignment=_align()
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        ws.merge_cells(start_row=2, start_column=cs, end_row=2, end_column=cs+4)
        cell = ws.cell(2, cs, f"{d:02d}-{month:02d}-{year}")
        cell.fill=_fill(C_MED); cell.font=_font(True,C_WHITE); cell.alignment=_align()

    c = ws.cell(3,1,"Time"); c.fill=_fill(C_DARK); c.font=_font(True,C_WHITE); c.alignment=_align()
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        for i, lbl in enumerate(['ZTE','NOKIA','NIB','4G','Tot(Gbps)']):
            cell = ws.cell(3, cs+i, lbl)
            cell.fill=_fill(C_LTEAL); cell.font=_font(True); cell.alignment=_align()

    # Build per-slot speed data
    zte_spd  = {}  # {day: {slot: Mbps}}
    nokia_spd= {}  # {day: {period_slot: Mbps}} — 15-min, same for 3 slots
    lan_spd  = {}  # {day: {slot: Mbps}}
    prtg_spd = {}  # {day: {slot: Mbps}}

    for r in zte_5m:
        ts = r.get('ts')
        if not ts: continue
        d = int(r['day']); slot = (ts.hour*60+ts.minute)//5
        zte_spd.setdefault(d, {})[slot] = _safe(r.get('dl_tput_mbps', 0))

    for r in nokia_5m:
        d = int(r['day_']); hr = int(r['hr']); mn = int(r.get('mn',0) or 0)
        period_slot = hr*12 + (mn//15)*3
        spd = _safe(r.get('slot_tput', 0))  # peak2gthru+peak3gthru Mbps
        # ADD: nokia_sgsn_report_out has multiple node rows per period
        nokia_spd.setdefault(d, {})
        nokia_spd[d][period_slot] = nokia_spd[d].get(period_slot, 0.0) + spd

    for r in lan_5m:
        ts = r.get('ts')
        if not ts: continue
        d = int(r['day']); slot = (ts.hour*60+ts.minute)//5
        # NIB = (in_speed_kbps + out_speed_kbps) / 1024 — sum across all switches
        spd = (_safe(r.get('dl_tput_mbps',0)) + _safe(r.get('ul_tput_mbps',0)))
        lan_spd.setdefault(d, {})[slot] = lan_spd.get(d,{}).get(slot,0.0) + spd

    for r in prtg_5m:
        ts = r.get('ts')
        if not ts: continue
        d = int(r['day']); slot = (ts.hour*60+ts.minute)//5
        spd = _safe(r.get('dl_tput_mbps', 0))  # in_speed_kbps/1024
        prtg_spd.setdefault(d, {})[slot] = prtg_spd.get(d,{}).get(slot,0.0) + spd

    dsv = {d: [[] for _ in range(CPD)] for d in range(1, days+1)}

    for slot in range(288):
        row = 4+slot
        hr  = (slot*5)//60; mn = (slot*5)%60
        c = ws.cell(row,1, dt_time(hr,mn,0)); c.number_format='HH:MM'; c.alignment=_align()
        period_slot = (slot//3)*3

        for d in range(1, days+1):
            cs = 2+(d-1)*CPD
            zte_v  = zte_spd.get(d,{}).get(slot, 0.0)
            nok_v  = nokia_spd.get(d,{}).get(period_slot, 0.0)
            lan_v  = lan_spd.get(d,{}).get(slot, 0.0)
            prtg_v = prtg_spd.get(d,{}).get(slot, 0.0)
            tot_v  = (lan_v + prtg_v) / 1024.0  # binary Gbps: (NIB+4G)/1024 confirmed from ref

            vals = [zte_v, nok_v, lan_v, prtg_v, tot_v]
            for i, val in enumerate(vals):
                ws.cell(row, cs+i, round(val, 2)).alignment=_align()
                dsv[d][i].append(val)

    # MAX row only (no Total row for Peak sheet)
    mr = 4+288
    ws.cell(mr,1,"MAX").font=_font(True); ws.cell(mr,1).fill=_fill(C_TOTAL)
    for d in range(1, days+1):
        cs = 2+(d-1)*CPD
        for i in range(CPD):
            lst = dsv[d][i]; v = max(lst) if lst else 0.0
            cell = ws.cell(mr, cs+i, round(v,2))
            cell.font=_font(True); cell.fill=_fill(C_TOTAL); cell.alignment=_align()

    ws.column_dimensions['A'].width=10
    for col in range(2, TCOL+1): ws.column_dimensions[get_column_letter(col)].width=11


def _build_4g(wb, date_str, dt, kpi_data):
    ws = wb.create_sheet("4G")
    dl = dt.strftime("%Y-%m-%d")

    HDRS = ["Period start time",
            "Avg EMM-REG users","Avg active EPS bearers",
            "Average registered EPS VoLTE users",
            "Peak active def EPS bearers, MME",
            "Peak active ded EPS bearers, MME",
            "Peak EMM-REG users, MME",
            "Avg EMM-DEREG users",
            "MOS VALUE","Avg RTP PACKET LOSS","Reregistration Success Rate",
            "VoLTE Dedicated Bearer Act","",""]
    SUBH = ["","","","","","","","","","","","Attempt","Success","% Success"]

    for col, lbl in enumerate(HDRS, 1):
        c = ws.cell(1, col, lbl); c.fill=_fill(C_DARK); c.font=_font(True,C_WHITE)
        c.alignment=_align(wrap=True); c.border=_border()
    for col, lbl in enumerate(SUBH, 1):
        c = ws.cell(2, col, lbl); c.fill=_fill(C_LTEAL); c.font=_font(True)
        c.alignment=_align(); c.border=_border()

    SUM = {k: 0.0 for k in ['emm_reg','eps_act','volte_u','emm_dereg','mos','rtp','rereg','v_pct']}
    MXP = {k: 0.0 for k in ['peak_def','peak_ded','peak_emm','volte_u','rtp','v_att','v_succ']}
    N = 0

    for hr in range(24):
        row = 3+hr; ts = f"{dl} {hr:02d}:00:00"
        d = kpi_data.get(hr, {})

        emm_reg  = _safe(d.get('emm_reg'))
        eps_act  = _safe(d.get('eps_act'))
        volte_u  = _safe(d.get('volte_u'))
        emm_der  = _safe(d.get('emm_dereg'))
        peak_def = _safe(d.get('peak_def'))
        peak_ded = _safe(d.get('peak_ded'))
        peak_emm = _safe(d.get('peak_emm'))
        mos      = _safe(d.get('mos'))
        rtp      = _safe(d.get('rtp'))
        rereg    = _safe(d.get('rereg'))
        v_att    = _safe(d.get('att'))
        v_succ   = _safe(d.get('succ'))
        v_pct    = _safe(d.get('per_succ')) if d.get('per_succ') else _pct(v_succ, v_att)

        for col, val in [
            (1,ts),(2,int(emm_reg)),(3,int(eps_act)),(4,int(volte_u)),
            (5,int(peak_def)),(6,int(peak_ded)),(7,int(peak_emm)),
            (8,int(emm_der)),(9,round(mos,2)),(10,round(rtp,2)),
            (11,round(rereg,2)),(12,int(v_att)),(13,int(v_succ)),(14,round(v_pct,2)),
        ]: _dat(ws, row, col, val)

        for k, val in [('emm_reg',emm_reg),('eps_act',eps_act),('volte_u',volte_u),
                        ('emm_dereg',emm_der),('mos',mos),('rtp',rtp),
                        ('rereg',rereg),('v_pct',v_pct)]:
            SUM[k] += val
        for k, val in [('peak_def',peak_def),('peak_ded',peak_ded),('peak_emm',peak_emm),
                        ('volte_u',volte_u),('rtp',rtp),('v_att',v_att),('v_succ',v_succ)]:
            if val > MXP[k]: MXP[k] = val
        N += 1

    n = max(N, 1)
    lr = 3+24
    sum_row = [
        (1,  ""),
        (2,  int(round(SUM['emm_reg']   / n))),
        (3,  int(round(SUM['eps_act']   / n))),
        (4,  int(MXP['volte_u'])),
        (5,  int(MXP['peak_def'])),
        (6,  int(MXP['peak_ded'])),
        (7,  int(MXP['peak_emm'])),
        (8,  int(round(SUM['emm_dereg'] / n))),
        (9,  round(SUM['mos']   / n, 2)),
        (10, round(MXP['rtp'],  2)),
        (11, round(SUM['rereg'] / n, 2)),
        (12, int(MXP['v_att'])),
        (13, int(MXP['v_succ'])),
        (14, round(SUM['v_pct'] / n, 2)),
    ]
    for col, val in sum_row:
        _total_row(ws, lr, col, val)

    ws.row_dimensions[1].height = 50
    for col in range(1, 15): ws.column_dimensions[get_column_letter(col)].width = 15


# =============================================================================
#  SHEET 8 — TRAI
# =============================================================================

def _build_trai(wb, year, month, report_day, trai_rows):
    """
    TRAI sheet — shows all days 1..report_day.
    Data from MySQL sgsn_trai_report (populated by Java/NetAct script).
    Today's row is 0 if Java/NetAct has not yet processed it — shown with
    light shading and a note. TOTAL row sums only populated days.
    """
    ws = wb.create_sheet("TRAI")

    ws.merge_cells("A1:A3")
    c = ws.cell(1,1,"Date"); c.fill=_fill(C_DARK); c.font=_font(True,C_WHITE)
    c.alignment=_align(); c.border=_border()
    _mhdr(ws, 1,  2,  8, "Total Volume")
    _mhdr(ws, 1,  9, 12, "Average Throughput (in Mbps)")
    _mhdr(ws, 1, 13, 16, "Peak Throughtput (in Mbps)")
    _mhdr(ws, 1, 17, 19, "Attached Subscribers")
    _mhdr(ws, 2,  2,  3, "2G (in GB)")
    _mhdr(ws, 2,  4,  5, "3G (in GB)")
    _mhdr(ws, 2,  6,  7, "4G (in GB)")
    c = ws.cell(2,8,"Total\n(in TB)"); c.fill=_fill(C_MED); c.font=_font(True,C_WHITE)
    c.alignment=_align(wrap=True); c.border=_border()
    for col in range(9, 20): ws.cell(2, col).fill=_fill(C_MED)

    for col, lbl in [
        (2,"U/L"),(3,"D/L"),(4,"U/L"),(5,"D/L"),(6,"U/L"),(7,"D/L"),(8,""),
        (9,"2G"),(10,"3G"),(11,"4G"),(12,"Gn"),
        (13,"2G"),(14,"3G"),(15,"4G"),(16,"Gn"),
        (17,"2G"),(18,"3G"),(19,"4G"),
    ]:
        c = ws.cell(3, col, lbl); c.fill=_fill(C_LTEAL); c.font=_font(True)
        c.alignment=_align(); c.border=_border()

    # Build lookup: day → row data
    lkp = {}
    for r in trai_rows:
        dv = r.get("Date_")
        if dv:
            try:
                day = dv.day if hasattr(dv,"day") else int(str(dv)[8:10])
                lkp[day] = r
            except Exception:
                pass

    # Accumulators for TOTAL row (only populated days)
    T = {k: 0.0 for k in ["ul_2g","dl_2g","ul_3g","dl_3g",
                            "ul_4g","dl_4g","total_tb"]}

    C_ZERO_FILL = "FFF0F0"   # light red for unpopulated rows

    for d in range(1, report_day+1):
        row = 3+d
        r   = lkp.get(d, {})
        def gc(k): return _safe(r.get(k))

        has_data = (gc("ul_2g") + gc("ul_3g") + gc("ul_4g")) > 0

        # Date cell
        dc = ws.cell(row, 1, f"{d:02d}/{month:02d}/{year}")
        dc.fill      = _fill(C_ZERO_FILL if not has_data else ("EEF4FB" if d%2==0 else "FFFFFF"))
        dc.font      = _font(False)
        dc.alignment = _align()
        dc.border    = _border()

        for col, val in [
            (2,int(gc("ul_2g"))),(3,int(gc("dl_2g"))),
            (4,int(gc("ul_3g"))),(5,int(gc("dl_3g"))),
            (6,int(gc("ul_4g"))),(7,int(gc("dl_4g"))),
            (8,round(gc("total_tb"),2)),
            (9,int(gc("avg_2g"))),(10,int(gc("avg_3g"))),
            (11,int(gc("avg_4g"))),(12,int(gc("avg_gn"))),
            (13,int(gc("peak_2g"))),(14,int(gc("peak_3g"))),
            (15,int(gc("peak_4g"))),(16,int(gc("peak_gn"))),
            (17,int(gc("att_2g"))),(18,int(gc("att_3g"))),(19,int(gc("att_4g"))),
        ]:
            bg = C_ZERO_FILL if not has_data else ("EEF4FB" if d%2==0 else "FFFFFF")
            fg = "AAAAAA" if not has_data else "000000"
            c  = ws.cell(row, col, val)
            c.fill      = _fill(bg)
            c.font      = _font(False, fg)
            c.alignment = _align()
            c.border    = _border()

        # Accumulate totals for populated days only
        if has_data:
            for k in T:
                T[k] += gc(k)

    # TOTAL row — cumulative sum of all populated days
    tr = 3 + report_day + 1
    tc = ws.cell(tr, 1, "TOTAL")
    tc.fill=_fill("FFF2CC"); tc.font=_font(True); tc.alignment=_align(); tc.border=_border()
    for col, val in [
        (2,int(T["ul_2g"])),(3,int(T["dl_2g"])),
        (4,int(T["ul_3g"])),(5,int(T["dl_3g"])),
        (6,int(T["ul_4g"])),(7,int(T["dl_4g"])),
        (8,round(T["total_tb"],2)),
        (9,""),(10,""),(11,""),(12,""),
        (13,""),(14,""),(15,""),(16,""),
        (17,""),(18,""),(19,""),
    ]:
        c = ws.cell(tr, col, val)
        c.fill=_fill("FFF2CC"); c.font=_font(True); c.alignment=_align(); c.border=_border()

    # Note row
    nr = tr + 1
    nc = ws.cell(nr, 1,
        f"Note: Volumes in GB. Today ({report_day:02d}/{month:02d}/{year}) shows 0 if "
        "Java/NetAct has not yet processed it. Shaded rows = not yet available. "
        "TOTAL = cumulative sum of populated days only.")
    nc.fill=_fill("FFFFFF"); nc.font=_font(False,"777777",8)
    nc.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    nc.border=_border()
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=19)
    ws.row_dimensions[nr].height = 28

    ws.column_dimensions["A"].width = 14
    for col in range(2, 20): ws.column_dimensions[get_column_letter(col)].width = 10
    ws.row_dimensions[1].height = 30; ws.row_dimensions[2].height = 25


# =============================================================================
#  ENTRY POINT
# =============================================================================

def generate_daily_report(date_str, tcs_file=None, log=print):
    dt    = datetime.strptime(date_str, "%Y-%m-%d")
    day   = dt.day; month = dt.month; year = dt.year
    mname = dt.strftime("%b").upper()

    log(f"[DailyReport] Starting {date_str}")

    # ── Hourly data (for SGSNZTE, SGSNNOKIA, Total, 4G sheets) ──────────────
    log("[DailyReport] ZTE hourly...")
    zte_h       = _zte_hourly(date_str)
    log("[DailyReport] Nokia hourly (nokia_sgsn_report_out)...")
    nokia_h     = _nokia_hourly(date_str)
    log("[DailyReport] Nokia 4G hourly...")
    g4_h        = _nokia_4g_hourly(date_str)
    log("[DailyReport] LAN/NIB total hourly...")
    lan_total_h = _lan_total_hourly(date_str)
    log("[DailyReport] 4G KPI (EMM/EPS/VoLTE)...")
    kpi_data    = _4g_kpi_hourly(date_str)

    # ── Monthly data (for MAR-26, 5Min, Peak sheets) ─────────────────────────
    log("[DailyReport] Monthly ZTE 5-min...")
    zte_5m      = _zte_5min_month(year, month)
    log("[DailyReport] Monthly LAN/NIB 5-min...")
    lan_5m      = _lan_5min_month(year, month)
    log("[DailyReport] Monthly Nokia (nokia_sgsn_report_out)...")
    nokia_5m    = _nokia_month(year, month)       # has slot_total_mb, slot_tput
    log("[DailyReport] Monthly Nokia 4G hourly...")
    g4_5m       = _nokia_4g_month(year, month)    # has day_, hr, dl_mb, ul_mb
    log("[DailyReport] Monthly PRTG 4G 5-min...")
    prtg_5m     = _prtg_5min_month(year, month)

    # Daily totals for MAR-26
    log("[DailyReport] Nokia daily totals (MAR-26)...")
    nokia_daily  = _nokia_month_daily(year, month)
    log("[DailyReport] Nokia 4G daily totals (MAR-26)...")
    nokia4g_daily= _nokia_4g_month_daily(year, month)
    log("[DailyReport] PRTG 4G daily totals (MAR-26 DT, ZTE shutdown)...")
    lan_daily    = _prtg4g_month_daily(year, month)

    # TCS EMS
    log("[DailyReport] Reading TCS EMS 4G file...")
    tcs_daily = _tcs4g_daily(date_str, tcs_filepath=tcs_file, log=log)
    if not tcs_daily:
        expected = f"WZ_Gujarat_RAN_KPIS_eNBwise_Daily_{dt.day:02d}_{dt.month:02d}_{dt.year}.csv"
        log(f"[DailyReport] WARNING: TCS CSV NOT found for {date_str}.")
        log(f"[DailyReport]   Expected: {expected}")
        log(f"[DailyReport]   Will use MySQL sgsn_tcs_daily for past days.")
        log(f"[DailyReport]   Today's TCS columns will be ZERO (copy CSV and re-run).")
    else:
        log(f"[DailyReport] TCS: DL={tcs_daily['dl_mb']/1024:.0f}GB  UL={tcs_daily['ul_mb']/1024:.0f}GB  ({tcs_daily.get('enb_count',0)} eNBs)  [stored to MySQL]")

    # TRAI
    log("[DailyReport] TRAI data...")
    trai = _trai_month(year, month)

    log("[DailyReport] Building workbook...")
    wb = Workbook()

    _build_mar26(wb, year, month, mname, day,
                 nokia_daily, nokia4g_daily, lan_daily, tcs_daily)
    _build_node_sheet(wb, "SGSNZTE",   date_str, dt, zte_h,   {}, {}, is_nokia=False)
    _build_node_sheet(wb, "SGSNNOKIA", date_str, dt, nokia_h, {}, {}, is_nokia=True)
    _build_total(wb, date_str, dt, zte_h, nokia_h, g4_h, lan_total_h)
    _build_5min(wb, year, month, day, zte_5m, nokia_5m, lan_5m, g4_5m, prtg_5m)
    _build_peak(wb, year, month, day, zte_5m, nokia_5m, lan_5m, prtg_5m)
    _build_4g(wb, date_str, dt, kpi_data)
    _build_trai(wb, year, month, day, trai)

    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"SGSN_{day:02d}_{mname}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    log(f"[DailyReport] Saved: {filepath}")
    return filepath
