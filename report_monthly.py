# report_monthly.py
# Monthly SGSN Summary Report
# Reads from MySQL: downldSgsn (GB volumes) + peakThroughputSgsn (Mbps)
# ALL volumes displayed in GB for reader convenience.

import calendar, os
from db import my_query
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR

# ── colour palette ────────────────────────────────────────────────────────
C_TITLE  = "1F3864"
C_HDR1   = "2E75B6"
C_HDR2   = "1F4E79"
C_ALT    = "EEF4FB"
C_WHITE  = "FFFFFF"
C_TOTAL  = "FFF2CC"
C_MAX    = "E2EFDA"
C_BORDER = "B8CCE4"

def _bord():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val, bg=C_WHITE, bold=False, sz=10,
          fg="000000", wrap=False, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=sz, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _bord()
    if num_fmt:
        c.number_format = num_fmt
    return c

def _merge(ws, row, c1, c2, val, bg=C_TITLE, sz=12, fg=C_WHITE):
    _cell(ws, row, c1, val, bg=bg, bold=True, sz=sz, fg=fg)
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row,   end_column=c2)


def generate_monthly_report(year, month, log=print):
    """
    Generate monthly SGSN report. All volumes in GB. Peak speeds in Mbps.
    Layout:
      Col A    : Date
      Col B    : Nokia 2G/3G (GB)
      Col C    : DT / LAN (GB)
      Col D    : Nokia 4G (GB)
      Col E    : TCS 4G (GB)   ← from sgsn_tcs_daily
      Col F    : ZTE 2G/3G (GB)
      Col G    : Grand Total (GB)
      Col H    : LAN UL peak (Mbps)
      Col I    : LAN DL peak (Mbps)
      Col J    : 4G UL peak (Mbps)
      Col K    : 4G DL peak (Mbps)
      Col L    : Nokia UL peak (Mbps)
      Col M    : Nokia DL peak (Mbps)
      Col N    : LAN+4G Total peak (Gbps)
    """
    month_name = calendar.month_name[month]
    log(f"[Monthly Report] Generating {month_name} {year}...")

    # ── Load volumes from downldSgsn (all GB) ────────────────────────────
    vol_rows = my_query("""
        SELECT Date_,
               downldSGSNZTE    AS zte_gb,
               downldSGSNdt     AS dt_gb,
               downldSGSNNOKIA  AS nokia_gb,
               download4g       AS g4_gb,
               downldSGSNtot    AS tot_gb
        FROM downldSgsn
        WHERE MONTH(Date_) = %s AND YEAR(Date_) = %s
        ORDER BY Date_
    """, (month, year))

    if not vol_rows:
        raise ValueError(
            f"No data in MySQL downldSgsn for {month_name} {year}.\n"
            f"Run 'Download Summary' for each day first."
        )

    # ── Load TCS from sgsn_tcs_daily ─────────────────────────────────────
    tcs_rows = my_query("""
        SELECT DAY(date) AS day_, dl_mb, ul_mb
        FROM sgsn_tcs_daily
        WHERE MONTH(date) = %s AND YEAR(date) = %s
    """, (month, year))
    tcs_dict = {}
    for r in tcs_rows:
        d = int(r["day_"])
        tcs_dict[d] = round((float(r["dl_mb"] or 0) + float(r["ul_mb"] or 0)) / 1024.0, 4)

    # ── Load peak throughput ──────────────────────────────────────────────
    peak_rows = my_query("""
        SELECT Date_,
               upPeakThruDt      AS lan_up,
               downPeakThruDt    AS lan_dn,
               upPeakThruNokia4g AS g4_up,
               downPeakThruNokia4g AS g4_dn,
               upPeakThruNokia   AS nok_up,
               downPeakThruNokia AS nok_dn,
               upPeakThruZte     AS zte_up,
               downPeakThruZte   AS zte_dn
        FROM peakThroughputSgsn
        WHERE MONTH(Date_) = %s AND YEAR(Date_) = %s
        ORDER BY Date_
    """, (month, year))
    peak_dict = {str(r["Date_"]): r for r in peak_rows}

    log(f"[Monthly Report] {len(vol_rows)} days loaded.")

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name[:3]}-{year}"
    ws.sheet_view.showGridLines = False

    NCOLS = 14

    # Row 1 — main title
    _merge(ws, 1, 1, NCOLS,
           f"BSNL Gujarat NOC  —  SGSN Monthly Traffic Report  —  {month_name} {year}",
           bg=C_TITLE, sz=13)
    ws.row_dimensions[1].height = 22

    # Row 2 — section sub-headers
    _merge(ws, 2, 1, 1,  "Date",           bg=C_HDR2, sz=10)
    _merge(ws, 2, 2, 7,  "Volume (GB)",    bg=C_HDR1, sz=10)
    _merge(ws, 2, 8, 14, "Peak Throughput", bg=C_HDR2, sz=10)
    ws.row_dimensions[2].height = 18

    # Row 3 — column headers
    hdrs = [
        "Date",
        "Nokia\n2G/3G (GB)", "DT / LAN\n(GB)", "Nokia\n4G (GB)",
        "TCS\n4G (GB)",      "ZTE\n2G/3G (GB)", "Grand\nTotal (GB)",
        "LAN UL\n(Mbps)",    "LAN DL\n(Mbps)",
        "4G UL\n(Mbps)",     "4G DL\n(Mbps)",
        "Nokia UL\n(Mbps)",  "Nokia DL\n(Mbps)",
        "LAN+4G\n(Gbps)",
    ]
    for ci, h in enumerate(hdrs, 1):
        _cell(ws, 3, ci, h, bg=C_HDR1, bold=True, fg=C_WHITE, sz=9, wrap=True, align="center")
    ws.row_dimensions[3].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────
    def fv(row, key):
        return round(float(row.get(key) or 0), 3)

    def fp(pk, key):
        if not pk: return 0.0
        return round(float(pk.get(key) or 0), 2)

    tot_nokia = tot_lan = tot_g4 = tot_tcs = tot_zte = tot_total = 0.0
    max_lan_up = max_lan_dn = max_g4_up = max_g4_dn = 0.0
    max_nok_up = max_nok_dn = max_gbps = 0.0

    FMT_GB   = "#,##0.000"
    FMT_MBPS = "#,##0.00"
    FMT_GBPS = "#,##0.0000"

    for i, r in enumerate(vol_rows):
        rn  = 4 + i
        bg  = C_ALT if i % 2 == 1 else C_WHITE
        dk  = str(r["Date_"])

        # import day number for TCS lookup
        from datetime import datetime
        try:
            day_num = datetime.strptime(dk.split()[0], "%Y-%m-%d").day
        except Exception:
            day_num = 0

        nokia_gb = fv(r, "nokia_gb")
        dt_gb    = fv(r, "dt_gb")
        g4_gb    = fv(r, "g4_gb")
        tcs_gb   = tcs_dict.get(day_num, 0.0)
        zte_gb   = fv(r, "zte_gb")
        # recalculate total including TCS
        total_gb = round(nokia_gb + dt_gb  + g4_gb + tcs_gb + zte_gb, 3)

        pk = peak_dict.get(dk, {})
        lan_up  = fp(pk, "lan_up");  lan_dn  = fp(pk, "lan_dn")
        g4_up   = fp(pk, "g4_up");   g4_dn   = fp(pk, "g4_dn")
        nok_up  = fp(pk, "nok_up");  nok_dn  = fp(pk, "nok_dn")
        tot_gbps = round((lan_up + lan_dn + g4_up + g4_dn) / 1024.0, 4)

        vals = [dk, nokia_gb, dt_gb,  g4_gb, tcs_gb, zte_gb, total_gb,
                lan_up, lan_dn, g4_up, g4_dn, nok_up, nok_dn, tot_gbps]
        fmts = [None, FMT_GB, FMT_GB, FMT_GB, FMT_GB, FMT_GB, FMT_GB,
                FMT_MBPS, FMT_MBPS, FMT_MBPS, FMT_MBPS, FMT_MBPS, FMT_MBPS, FMT_GBPS]

        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            al = "left" if ci == 1 else "center"
            _cell(ws, rn, ci, val, bg=bg, sz=10, align=al, num_fmt=fmt)

        # accumulate
        tot_nokia += nokia_gb; tot_lan += dt_gb;  tot_g4 += g4_gb
        tot_tcs   += tcs_gb;   tot_zte += zte_gb; tot_total += total_gb
        for var, cur in [(max_lan_up,lan_up),(max_lan_dn,lan_dn),
                         (max_g4_up,g4_up),(max_g4_dn,g4_dn),
                         (max_nok_up,nok_up),(max_nok_dn,nok_dn),
                         (max_gbps,tot_gbps)]:
            pass  # handled below
        if lan_up   > max_lan_up:  max_lan_up  = lan_up
        if lan_dn   > max_lan_dn:  max_lan_dn  = lan_dn
        if g4_up    > max_g4_up:   max_g4_up   = g4_up
        if g4_dn    > max_g4_dn:   max_g4_dn   = g4_dn
        if nok_up   > max_nok_up:  max_nok_up  = nok_up
        if nok_dn   > max_nok_dn:  max_nok_dn  = nok_dn
        if tot_gbps > max_gbps:    max_gbps    = tot_gbps

    last = 4 + len(vol_rows)

    # ── TOTAL row ──────────────────────────────────────────────────────────
    tot_vals = ["TOTAL",
                round(tot_nokia,3), round(tot_lan,3), round(tot_g4,3),
                round(tot_tcs,3),   round(tot_zte,3), round(tot_total,3),
                "", "", "", "", "", "", ""]
    tot_fmts = [None, FMT_GB, FMT_GB, FMT_GB, FMT_GB, FMT_GB, FMT_GB,
                None, None, None, None, None, None, None]
    for ci, (val, fmt) in enumerate(zip(tot_vals, tot_fmts), 1):
        _cell(ws, last, ci, val, bg=C_TOTAL, bold=True, sz=10,
              align="left" if ci==1 else "center", num_fmt=fmt)

    # ── MAX row ────────────────────────────────────────────────────────────
    max_vals = ["PEAK MAX",
                "", "", "", "", "", "",
                round(max_lan_up,2), round(max_lan_dn,2),
                round(max_g4_up,2),  round(max_g4_dn,2),
                round(max_nok_up,2), round(max_nok_dn,2),
                round(max_gbps,4)]
    max_fmts = [None, None, None, None, None, None, None,
                FMT_MBPS, FMT_MBPS, FMT_MBPS, FMT_MBPS,
                FMT_MBPS, FMT_MBPS, FMT_GBPS]
    for ci, (val, fmt) in enumerate(zip(max_vals, max_fmts), 1):
        _cell(ws, last+1, ci, val, bg=C_MAX, bold=True, sz=10,
              align="left" if ci==1 else "center", num_fmt=fmt)

    # ── Note row ──────────────────────────────────────────────────────────
    note_row = last + 2
    _cell(ws, note_row, 1,
          "Note: Volumes in GB. Nokia 2G/3G < DT/LAN < Nokia 4G < TCS 4G (expected — different traffic types).",
          bg=C_WHITE, sz=9, fg="555555", align="left")
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row,   end_column=NCOLS)

    # ── Column widths + freeze ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 13
    for ci in range(2, NCOLS+1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = "B4"

    # ── Save ──────────────────────────────────────────────────────────────
    fname = f"Monthly_SGSN_{year}_{month:02d}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)
    wb.save(path)
    log(f"[Monthly Report] Saved: {path}")
    return path
