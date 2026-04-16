# excel_helper.py
# Excel generation helpers using openpyxl

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
import os
from config import OUTPUT_DIR

# Color palette
C_HEADER_BG  = "1F3864"   # Dark blue header
C_HEADER_FG  = "FFFFFF"   # White text
C_SUB_BG     = "2E75B6"   # Section subheader
C_ROW_ALT    = "EEF4FB"   # Alternate row light blue
C_ROW_WHITE  = "FFFFFF"
C_BORDER     = "B8CCE4"
C_TOTAL_BG   = "FFF2CC"   # Yellow for totals
C_TITLE_BG   = "1F3864"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr_font(size=11, bold=True, color=C_HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)


def body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)


def hdr_fill(color=C_HEADER_BG):
    return PatternFill("solid", fgColor=color)


def row_fill(alt=False):
    return PatternFill("solid", fgColor=C_ROW_ALT if alt else C_ROW_WHITE)


def total_fill():
    return PatternFill("solid", fgColor=C_TOTAL_BG)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def right():
    return Alignment(horizontal="right", vertical="center")


def left_align():
    return Alignment(horizontal="left", vertical="center")


def write_title(ws, row, col, text, span=1, bg=C_HEADER_BG, fg=C_HEADER_FG, size=12):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", size=size, bold=True, color=fg)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = center()
    cell.border = BORDER
    if span > 1:
        end_col = get_column_letter(col + span - 1)
        ws.merge_cells(f"{get_column_letter(col)}{row}:{end_col}{row}")


def write_header(ws, row, headers, col_start=1):
    """Write a single header row with dark blue background."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = hdr_font()
        c.fill = hdr_fill()
        c.alignment = center()
        c.border = BORDER


def write_subheader(ws, row, headers, col_start=1):
    """Write a subheader row with medium blue background."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = hdr_fill(C_SUB_BG)
        c.alignment = center()
        c.border = BORDER


def write_row(ws, row, values, alt=False, bold=False, is_total=False, col_start=1):
    """Write a data row with alternating colors."""
    fill = total_fill() if is_total else row_fill(alt)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=col_start + i, value=v)
        c.font = body_font(bold=bold)
        c.fill = fill
        c.alignment = center() if i > 0 else left_align()
        c.border = BORDER


def set_col_widths(ws, widths):
    """Set column widths. widths is list of (col_index, width) tuples."""
    for col_idx, w in widths:
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def freeze(ws, row=2, col=1):
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def num_fmt(ws, row, col, fmt="#,##0.00"):
    ws.cell(row=row, column=col).number_format = fmt


def output_path(filename):
    return os.path.join(OUTPUT_DIR, filename)


def save_wb(wb, filename):
    path = output_path(filename)
    wb.save(path)
    return path
