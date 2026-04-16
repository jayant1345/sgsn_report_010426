# report_trai.py
# TRAI Regulatory Monthly Report
# Reads from MySQL reportsDB.sgsn_trai_report (populated by Java/NetAct script)
#
# Column units confirmed from reference XLS:
#   vol_2g_up_in_mb  → stored as GB (column name misleading, actual data = GB)
#   vol_tot_in_gb    → stored as GB → /1024 = TB
#   avg/peak_*_tput  → Mbps
#   attach_sub_*     → subscriber count

import calendar, os
from db import my_query
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR

C_TITLE  = "1F3864"
C_HDR1   = "2E75B6"
C_HDR2   = "1F4E79"
C_ALT    = "EEF4FB"
C_WHITE  = "FFFFFF"
C_TOTAL  = "FFF2CC"
C_ZERO   = "FFF0F0"
C_BORDER = "B8CCE4"

def _bord():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _cell(ws, row, col, val, bg=C_WHITE, bold=False, sz=10,
          fg="000000", wrap=False, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name="Calibri", size=sz, bold=bold, color=fg)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _bord()
    if num_fmt:
        c.number_format = num_fmt
    return c

def _merge(ws, row, c1, c2, val, bg=C_TITLE, sz=12, fg=C_WHITE):
    _cell(ws, row, c1, val, bg=bg, bold=True, sz=sz, fg=fg)
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row,   end_column=c2)


def generate_trai_report(year, month, log=print):
    """
    Generate TRAI monthly report.
    Shows ALL days of the month with data from sgsn_trai_report.
    Days not yet populated by Java/NetAct show as 0 with shading.
    Includes cumulative TOTAL row and monthly MAX row.
    """
    month_name = calendar.month_name[month]
    log(f"[TRAI Report] Querying sgsn_trai_report for {month_name} {year}...")

    rows = my_query("""
        SELECT
            DATE_FORMAT(`date`, '%%d/%%m/%%Y')   AS date_fmt,
            DAY(`date`)                           AS day_num,
            round(`vol_2g_up_in_mb`, 0)           AS v2g_ul,
            round(`vol_2g_dn_in_mb`, 0)           AS v2g_dl,
            round(`vol_3g_up_in_mb`, 0)           AS v3g_ul,
            round(`vol_3g_dn_in_mb`, 0)           AS v3g_dl,
            round(`vol_4g_up_in_mb`, 0)           AS v4g_ul,
            round(`vol_4g_dn_in_mb`, 0)           AS v4g_dl,
            round(`vol_tot_in_gb` / 1024.0, 2)    AS tot_tb,
            round(`avg_2g_tput`, 0)               AS a2g,
            round(`avg_3g_tput`, 0)               AS a3g,
            round(`avg_4g_tput`, 0)               AS a4g,
            round(`avg_gn_tput`, 0)               AS agn,
            round(`peak_2g_tput`, 0)              AS p2g,
            round(`peak_3g_tput`, 0)              AS p3g,
            round(`peak_4g_tput`, 0)              AS p4g,
            round(`peak_gn_tput`, 0)              AS pgn,
            round(`attach_sub_2g`, 0)             AS att2g,
            round(`attach_sub_3g`, 0)             AS att3g,
            round(`attach_sub_4g`, 0)             AS att4g
        FROM `reportsDB`.`sgsn_trai_report`
        WHERE month(`date`) = %s AND year(`date`) = %s
        ORDER BY `date`
    """, (month, year))

    if not rows:
        raise ValueError(
            f"No TRAI data for {month_name} {year} in MySQL sgsn_trai_report.\n"
            f"This table is populated by the Java/NetAct script — "
            f"check that the script has run for this month."
        )

    log(f"[TRAI Report] {len(rows)} rows found.")

    # Count populated vs zero rows
    populated = [r for r in rows if float(r.get("v2g_ul") or 0) +
                 float(r.get("v3g_ul") or 0) + float(r.get("v4g_ul") or 0) > 0]
    zero_rows = [r for r in rows if r not in populated]
    log(f"[TRAI Report] Populated: {len(populated)} days  |  Zero/not yet: {len(zero_rows)} days")

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "TRAI"
    ws.sheet_view.showGridLines = False

    NCOLS = 19

    # Row 1 — main title
    _merge(ws, 1, 1, NCOLS,
           f"BSNL Gujarat NOC  —  TRAI SGSN Regulatory Report  —  {month_name} {year}",
           bg=C_TITLE, sz=13)
    ws.row_dimensions[1].height = 22

    # Row 2 — section headers
    _merge(ws, 2, 1, 1,  "Date",                       bg=C_HDR2, sz=10)
    _merge(ws, 2, 2, 8,  "Total Volume",                bg=C_HDR1, sz=10)
    _merge(ws, 2, 9, 12, "Average Throughput (in Mbps)",bg=C_HDR2, sz=10)
    _merge(ws, 2,13,16,  "Peak Throughput (in Mbps)",   bg=C_HDR1, sz=10)
    _merge(ws, 2,17,19,  "Attached Subscribers",        bg=C_HDR2, sz=10)
    ws.row_dimensions[2].height = 18

    # Row 3 — sub-section + unit headers
    _merge(ws, 3, 1, 1,  "",          bg=C_HDR2)
    _merge(ws, 3, 2, 3,  "2G (in GB)",bg=C_HDR1, sz=9)
    _merge(ws, 3, 4, 5,  "3G (in GB)",bg=C_HDR1, sz=9)
    _merge(ws, 3, 6, 7,  "4G (in GB)",bg=C_HDR1, sz=9)
    _merge(ws, 3, 8, 8,  "Total\n(in TB)", bg=C_HDR1, sz=9)
    for ci, lbl, bg in [
        (9,"Total",C_HDR2),(10,"",C_HDR2),(11,"",C_HDR2),(12,"",C_HDR2),
        (13,"Total",C_HDR1),(14,"",C_HDR1),(15,"",C_HDR1),(16,"",C_HDR1),
        (17,"Total",C_HDR2),(18,"",C_HDR2),(19,"",C_HDR2),
    ]:
        _cell(ws, 3, ci, lbl, bg=bg, bold=True, fg=C_WHITE, sz=9)
    ws.row_dimensions[3].height = 26

    # Row 4 — column sub-headers (U/L, D/L, 2G/3G/4G/Gn)
    sub_hdrs = [
        "Date",
        "U/L", "D/L", "U/L", "D/L", "U/L", "D/L", "",
        "2G", "3G", "4G", "Gn",
        "2G", "3G", "4G", "Gn",
        "2G", "3G", "4G"
    ]
    for ci, h in enumerate(sub_hdrs, 1):
        bg = C_HDR2 if ci in (1,9,10,11,12,17,18,19) else C_HDR1
        _cell(ws, 4, ci, h, bg=bg, bold=True, fg=C_WHITE, sz=9, align="center")
    ws.row_dimensions[4].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    def fv(r, k):
        return float(r.get(k) or 0)

    # accumulators for TOTAL row
    TOT = {k: 0.0 for k in ['v2g_ul','v2g_dl','v3g_ul','v3g_dl',
                              'v4g_ul','v4g_dl','tot_tb']}
    MAX = {k: 0.0 for k in ['a2g','a3g','a4g','agn',
                              'p2g','p3g','p4g','pgn',
                              'att2g','att3g','att4g']}

    FMT_GB   = "#,##0"
    FMT_TB   = "#,##0.00"
    FMT_MBPS = "#,##0"
    FMT_SUB  = "#,##0"

    for i, r in enumerate(rows):
        rn  = 5 + i
        # determine if this row has real data
        has_data = (fv(r,'v2g_ul') + fv(r,'v3g_ul') + fv(r,'v4g_ul')) > 0
        bg = C_ZERO if not has_data else (C_ALT if i % 2 == 1 else C_WHITE)

        v2g_ul = fv(r,'v2g_ul'); v2g_dl = fv(r,'v2g_dl')
        v3g_ul = fv(r,'v3g_ul'); v3g_dl = fv(r,'v3g_dl')
        v4g_ul = fv(r,'v4g_ul'); v4g_dl = fv(r,'v4g_dl')
        tot_tb = fv(r,'tot_tb')
        a2g=fv(r,'a2g'); a3g=fv(r,'a3g'); a4g=fv(r,'a4g'); agn=fv(r,'agn')
        p2g=fv(r,'p2g'); p3g=fv(r,'p3g'); p4g=fv(r,'p4g'); pgn=fv(r,'pgn')
        att2g=fv(r,'att2g'); att3g=fv(r,'att3g'); att4g=fv(r,'att4g')

        vals_fmts = [
            (r["date_fmt"], None),
            (v2g_ul, FMT_GB), (v2g_dl, FMT_GB),
            (v3g_ul, FMT_GB), (v3g_dl, FMT_GB),
            (v4g_ul, FMT_GB), (v4g_dl, FMT_GB),
            (tot_tb, FMT_TB),
            (a2g, FMT_MBPS), (a3g, FMT_MBPS), (a4g, FMT_MBPS), (agn, FMT_MBPS),
            (p2g, FMT_MBPS), (p3g, FMT_MBPS), (p4g, FMT_MBPS), (pgn, FMT_MBPS),
            (att2g, FMT_SUB), (att3g, FMT_SUB), (att4g, FMT_SUB),
        ]
        for ci, (val, fmt) in enumerate(vals_fmts, 1):
            al = "left" if ci == 1 else "center"
            fg = "AAAAAA" if (not has_data and ci > 1) else "000000"
            _cell(ws, rn, ci, val, bg=bg, sz=10, align=al, num_fmt=fmt, fg=fg)

        if has_data:
            for k in TOT:  TOT[k] += fv(r,k)
            for k in MAX:
                v = fv(r,k)
                if v > MAX[k]: MAX[k] = v

    last = 5 + len(rows)

    # ── TOTAL row (cumulative sum of populated days) ────────────────────
    _cell(ws, last, 1, "TOTAL", bg=C_TOTAL, bold=True, sz=10, align="left")
    tot_vals_fmts = [
        (round(TOT['v2g_ul'],0), FMT_GB),  (round(TOT['v2g_dl'],0), FMT_GB),
        (round(TOT['v3g_ul'],0), FMT_GB),  (round(TOT['v3g_dl'],0), FMT_GB),
        (round(TOT['v4g_ul'],0), FMT_GB),  (round(TOT['v4g_dl'],0), FMT_GB),
        (round(TOT['tot_tb'], 2), FMT_TB),
        ("",""),("",""),("",""),("",""),
        ("",""),("",""),("",""),("",""),
        ("",""),("",""),("",""),
    ]
    for ci, (val,fmt) in enumerate(tot_vals_fmts, 2):
        _cell(ws, last, ci, val, bg=C_TOTAL, bold=True, sz=10, num_fmt=fmt or None)

    # ── MAX row (peak values across all days) ──────────────────────────
    _cell(ws, last+1, 1, "PEAK MAX", bg="E2EFDA", bold=True, sz=10, align="left")
    max_vals_fmts = [
        ("",""),("",""),("",""),("",""),("",""),("",""),("",""),
        (round(MAX['a2g'],0),FMT_MBPS),(round(MAX['a3g'],0),FMT_MBPS),
        (round(MAX['a4g'],0),FMT_MBPS),(round(MAX['agn'],0),FMT_MBPS),
        (round(MAX['p2g'],0),FMT_MBPS),(round(MAX['p3g'],0),FMT_MBPS),
        (round(MAX['p4g'],0),FMT_MBPS),(round(MAX['pgn'],0),FMT_MBPS),
        (round(MAX['att2g'],0),FMT_SUB),(round(MAX['att3g'],0),FMT_SUB),
        (round(MAX['att4g'],0),FMT_SUB),
    ]
    for ci, (val,fmt) in enumerate(max_vals_fmts, 2):
        _cell(ws, last+1, ci, val, bg="E2EFDA", bold=True, sz=10, num_fmt=fmt or None)

    # ── Note rows ─────────────────────────────────────────────────────────
    note1 = last + 2
    _cell(ws, note1, 1,
          f"Data available for {len(populated)} of {len(rows)} days.  "
          f"Shaded rows (if any) = not yet populated by Java/NetAct script.",
          bg=C_WHITE, sz=9, fg="555555", align="left")
    ws.merge_cells(start_row=note1, start_column=1,
                   end_row=note1,   end_column=NCOLS)

    note2 = last + 3
    _cell(ws, note2, 1,
          "Volumes in GB. Total in TB. Throughput in Mbps. "
          "TOTAL row = cumulative sum of populated days. PEAK MAX = highest daily value.",
          bg=C_WHITE, sz=9, fg="555555", align="left")
    ws.merge_cells(start_row=note2, start_column=1,
                   end_row=note2,   end_column=NCOLS)

    # ── Column widths + freeze ─────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    for ci in range(2, NCOLS+1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "B5"

    # ── Save ──────────────────────────────────────────────────────────────
    fname = f"TRAI_SGSN_{year}_{month:02d}.xlsx"
    path  = os.path.join(OUTPUT_DIR, fname)
    wb.save(path)
    log(f"[TRAI Report] Saved: {path}  ({len(populated)} days populated)")
    return path
